
from openpyxl.styles import Font, Alignment, PatternFill, colors


from run8 import  *
from openpyxl import *


from openpyxl.utils import FORMULAE

thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))



class report:



    def __init__(self,root):

        self.list_original=create_all_list()[1] # c'est la list des titres de tableau de fichier score....
        self.list_report=create_all_list()[0] # c'est le global des titrs qui se trouve dans  le rapport final
        self.list_report2=create_all_list()[2] # c'est pour regler le dicalage ici ona utiliser list_state_titles2 unitile
        self.list_red=[] # list_red c'est l'esemble des noms qui ont couleur rouge

        self.root = root
        self.root.title('KPI Tool')
        self.interface()


    # lire les donnéés avec pandas

    def interface(self):
        self.frame1 = Frame(self.root, width=340, bg='#0000FF',height=300)
        self.frame1.grid(row=0, column=0, ipady=10, ipadx=10)
        self.frame1.grid_propagate(0)

        self.frame1.grid_rowconfigure(0, weight=3)
        self.frame1.grid_rowconfigure(1, weight=3)
        self.frame1.grid_rowconfigure(2, weight=3)
        self.frame1.grid_rowconfigure(3, weight=2)
        #self.frame1.grid_columnconfigure(0, weight=2)

        self.vfichier1 = StringVar()
        self.vfichier2 = StringVar()
        self.vfichier1.set('')
        self.vfichier2.set('')
        self.chemin = ''
        self.chemin1 = ''

        self.button1 = Button(self.frame1, text="ScoreCard File", command=self.set_fichier1,width=50,
                              height=2, bg='#66B239')
        self.button1.grid(row=0, column=0,columnspan=3, pady=5)

        self.button2 = Button(self.frame1, text='Follow-up File',command=self.set_fichier2, width=50, height=2,
                              bg='#66B239')
        self.button2.grid(row=1, column=0,columnspan=3, pady=5)

        self.button3 = Button(self.frame1, text='Generate report',command=self.set_emplacement, width=50, height=2,
                                 bg='#66B239')
        self.button3.grid(row=2, column=0, pady=5)
        self.progress_bar = Progressbar(self.frame1, orient='horizontal', length=286, mode='determinate')

    def set_fichier1(self):

        self.FILETYPES = [("text files", "*.xlsm;*.xlsx")]
        self.vfichier1.set(askopenfilename(filetypes=self.FILETYPES))
        if self.vfichier1.get() != '':
            self.button1['bg'] = '#006738'

    def set_fichier2(self):

        self.FILETYPES = [("text files", "*.xlsx")]
        self.vfichier2.set(askopenfilename(filetypes=self.FILETYPES))
        if self.vfichier2.get() != '':
            self.button2['bg'] = '#006738'

    def set_emplacement(self):
        import time
        self.FILETYPES = [("text files", "*.xlsx")]
        chemin1 = (askdirectory())
        date_now = time.strftime('%d%m%Y')
        def go():
            self.progress_bar.grid(row=3, column=0)

            self.progress_bar["value"] = 5
            self.root.update()

            self.red_items()  # c'est une list de range de liste self.list_names_titles qui [1...10]

            self.create_table()


            self.button1['bg'] = '#66B239'
            self.button2['bg'] = '#66B239'
            time.sleep(1)
            self.progress_bar.grid_forget()
            root.update()
            messagebox.showinfo(title=None,message="report successfully created")

            root.destroy()


        if chemin1 != '':
            self.chemin = chemin1 + '/' + 'F' + date_now + '.xlsx'
            my_file=Path( self.chemin)
            if my_file.exists():
                answer=messagebox.askquestion('file exists','The file already exists. Do you want to replace it? ')
                if answer=='yes':
                    go()


                else:None

            else:
                go()



    def read_data(self):
        import copy
        self.data1 = pd.read_excel(self.vfichier1.get(), sheet_name='VehicleScoreCard', skiprows=5, usecols='A : BQ',na_filter = False)
        self.data2 = pd.read_excel(self.vfichier2.get(), sheet_name='GLOBAL_INFO',usecols='A : BQ')
        self.data3 = pd.read_excel(self.vfichier2.get(), sheet_name='Retour_Analyse_eCLEM',usecols='A : CE')






        self.list_id_score=list(self.data1['ID'])
        self.list_mbd_score=list(self.data1['MBD'])
        self.list_SwQA_score=list(self.data1['SwQA'])

        self.list_service = list(self.data2['Service'])
        self.list_id = list(self.data2['ID CLEM O52_ASWP'])
        self.list_KPI1_c = list(self.data2['KPI1.c'])
        self.list_KPI1_d = list(self.data2['KPI1.d'])

        self.list_id_net=[]
        self.list_plus=[]
        for i in list(self.data1['ID']):
            if i !='nan':
                self.list_id_net.append(i)




    def insert_KPI(self,rapport,workbook): #inserer les données de KPI qui se trouvr dans le fichiers suivi.. sheet ..stratigie  dans la fonction reand data
        self.progress_bar["value"] = 60
        root.update()

        header_formattxt = Alignment(wrap_text=True)

        list_indice1=['KPI.1c\n'+i for i in list_indice]
        list_indice2=['KPI.1d\n'+i for i in list_indice]

        self.kpi_c=[]
        self.kpi_cv=[] #pour rapport2
        self.kpi_d=[]
        self.kpi_dv=[]
        kpic_instance=[]
        for i in range(len(self.list_report)):# l'objectif c'est trouver lindex de KPI dans le tableau de rapport pour
            # inserer les donné dans la coloune qui a ce index
            x = 6
            kpic_instance = []

            kpicv_instance = []# 2eme rapport

            kpid_instance = []

            kpidv_instance = [] # 2eme rapport

            if self.list_report[i] in list_indice1:


                for m, n in zip(self.list_KPI1_c, self.list_id):


                    x += 1
                    if n in list(self.data1['ID']):
                        v = list(self.data1['ID']).index(int(n)) + 7
                        k = "Only check rules eCLEM fill"
                        if k in str(m):

                            rapport.cell(v, i + 1).value = "Only check \n" \
                                                           "rules eCLEM \nfill"
                            rapport.cell(v, i + 1).border = thin_border
                            rapport.cell(v, i + 1).alignment = header_formattxt
                            rapport.cell(v, i + 1).fill = PatternFill(patternType='solid', fgColor=self.my_gray)
                            kpic_instance.append("Only check \n" \
                                                 "rules eCLEM \nfill")

                            kpicv_instance.append(v-7)

                        else:


                            #rapport.cell(v, i + 1).value = str(rapport.cell(v, i).value)
                            if str(rapport.cell(v, i ).value)=='0.00%' or str(rapport.cell(v, i ).value)=='0%':
                                 rapport.cell(v, i + 1).value = 'NA'
                                 rapport.cell(v, i + 1).fill = PatternFill(patternType='solid', fgColor=self.my_gray)
                                 kpic_instance.append('NA')

                            elif str(rapport.cell(v, i ).value)=='NO DATA':
                                rapport.cell(v, i + 1).value = 'NO DATA'
                                kpic_instance.append('NO DATA')

                            else:
                                rapport.cell(v, i + 1).value = str(rapport.cell(v, i).value)
                                rapport.cell(v, i + 1).fill = PatternFill(patternType='solid', fgColor=self.my_yallow)
                                kpic_instance.append(str(rapport.cell(v, i).value))
                            rapport.cell(v, i + 1).border = thin_border


                            kpicv_instance.append(v-7)

                        rapport.column_dimensions[get_column_letter(i + 1)].width = 15 # ici la dimmention des case de KPI


                self.kpi_c.append(kpic_instance)
                self.kpi_cv.append(kpicv_instance)





            if self.list_report[i] in list_indice2:#KPI.D
                for m, n in zip(self.list_KPI1_d, self.list_id):
                    x += 1
                    if n in list(self.data1['ID']):
                        v = list(self.data1['ID']).index(int(n)) + 7
                        k = "NA"

                        try:
                            b = int(rapport.cell(v, i - 1).value) + int(rapport.cell(v, i - 2).value) + int(
                                rapport.cell(v, i - 3).value)

                        except TypeError:
                            b=1
                        if k in str(m) and b!=0:

                            rapport.cell(v, i + 1).value = 'NA'
                            rapport.cell(v, i + 1).border = thin_border
                            rapport.cell(v, i + 1).alignment = header_formattxt
                            rapport.cell(v, i + 1).fill = PatternFill(patternType='solid', fgColor=self.my_gray)
                            kpidv_instance.append(v - 7)
                            kpid_instance.append('NA')



                        else:# insere kpi_d selon les condition


                            if (int(rapport.cell(v, i-7).value) - int(rapport.cell(v, i-6).value))!=0:

                                b=int(rapport.cell(v, i - 1).value) +int(rapport.cell(v, i - 2).value) +int(rapport.cell(v, i - 3).value)
                                a = int((int(rapport.cell(v, i - 3).value) * 1 + int(
                                    rapport.cell(v, i - 2).value) * 0.5 + int(
                                    rapport.cell(v, i - 1).value) * 0) * 100 / (
                                                int(rapport.cell(v, i - 7).value) - int(rapport.cell(v, i - 6).value)))

                                if a==0 and b!=0:

                                    rapport.cell(v, i + 1).value = 'NA'
                                    rapport.cell(v, i + 1).fill = PatternFill(patternType='solid', fgColor=self.my_gray)
                                    kpid_instance.append('NA')

                                elif b == 0 and str(rapport.cell(v, i - 4).value)!="Only check \nrules eCLEM \nfill" and str(rapport.cell(v, i - 4).value)!='NA' :
                                    rapport.cell(v, i + 1).value = 'NE'
                                    rapport.cell(v, i + 1).fill = PatternFill(patternType='solid', fgColor=self.my_white)
                                    kpid_instance.append('NE')

                                elif str(rapport.cell(v, i - 4).value)=="Only check \nrules eCLEM \nfill" or str(rapport.cell(v, i - 4).value)=='NA':
                                    rapport.cell(v, i + 1).value = 'NA'
                                    rapport.cell(v, i + 1).fill = PatternFill(patternType='solid',
                                                                              fgColor=self.my_gray)
                                    kpid_instance.append('NA')
                                elif b != 0 and str(rapport.cell(v, i - 4).value).strip() != "Only check \nrules eCLEM \nfill" and str(rapport.cell(v, i - 4).value) != 'NA':

                                    rapport.cell(v, i + 1).value = rapport.cell(v, i).value

                                    rapport.cell(v, i + 1).fill = PatternFill(patternType='solid',
                                                                              fgColor=self.my_yallow)
                                    kpid_instance.append(round(a, 2))

                                else:
                                    rapport.cell(v, i + 1).value = rapport.cell(v, i ).value
                                    rapport.cell(v, i + 1).fill = PatternFill(patternType='solid', fgColor=self.my_yallow)
                                    kpid_instance.append(round(a,2))






                            else:
                                    if b == 0 and str(rapport.cell(v, i - 4).value).strip()!="Only check \nrules eCLEM \nfill" and str(rapport.cell(v, i - 4).value)!='NA' :
                                            print(rapport.cell(v, i - 4).value)
                                            rapport.cell(v, i + 1).value = 'NE'
                                            rapport.cell(v, i + 1).fill = PatternFill(patternType='solid',
                                                                                      fgColor=self.my_white)
                                            kpid_instance.append('NE')
                                    elif str(rapport.cell(v, i - 4).value) == "Only check \nrules eCLEM \nfill" or str(
                                            rapport.cell(v, i - 4).value) != 'NA':
                                        rapport.cell(v, i + 1).value = 'NA'
                                        rapport.cell(v, i + 1).fill = PatternFill(patternType='solid', fgColor=self.my_gray)
                                        kpid_instance.append('NA')


                                    else:
                                        rapport.cell(v, i + 1).value = 'NA'
                                        rapport.cell(v, i + 1).fill = PatternFill(patternType='solid', fgColor=self.my_gray)
                                        kpid_instance.append('NA')


                            kpidv_instance.append(v - 7)
                            rapport.cell(v, i + 1).border = thin_border
                            cor='=Results_eCLEM!'+str(get_column_letter(i+1))+str(v)




                        rapport.column_dimensions[get_column_letter(i + 1)].width = 15
                self.kpi_d.append(kpid_instance)
                self.kpi_dv.append(kpidv_instance)


        self.state2(self.rapport2_id, workbook, self.kpi_raport2,self.kpi_c,self.kpi_cv,self.kpi_d,self.kpi_dv)

        go = report3(self.data1, self.data2,self.data3, workbook, self.list_red)


        #w = go.workbook

    def insert_data(self,rapport,workbook):# inserer les donneé de ID name , safty ,.. qui se trouve dans le fichier score
        self.progress_bar["value"] = 15
        root.update()

        self.idd=0

        self.rapport2_id=[]
        ##########################33

        header_formattxt = Alignment(wrap_text=True)

        for i in  range(len(self.list_report)):
            x = 6
            if self.list_report[i] in self.list_original and self.list_report[i] not in [' ','Unnamed','Service']:
                b = []
                for m in list(self.data1[self.list_report[i]]):

                    x += 1
                    if x<len( self.data1[self.list_report[i]])+7:# pour supprimer les case qui contient la valeur nan dans letableau score
                        rapport.cell(x, i + 1).value = m
                        b.append(m)
                        rapport.cell(x, i+1).border = thin_border
                self.rapport2_id.append(b)

            if self.list_report[i]=='Service':# inserer la colone service qui se trouve dans le fichier suivi ...
                b = []
                for m, n in zip(self.list_service, self.list_id):

                    x += 1
                    self.idd += 1
                    if n in list(self.data1['ID']):

                        v = list(self.data1['ID']).index(int(n)) + 7
                        rapport.cell(v, i+1).value = m

                        #b.append(self.list_service[v-1])

                        b.append(v-7)
                        rapport.cell(v, i+1).border = thin_border

                self.rapport2_id.append(b)

        self.state(rapport )
        self.insert_KPI(rapport,workbook)

        #self.state2(self.rapport2_id,workbook,[])


        ####################################################################################################################



        workbook.save(self.chemin)
        workbook.close()




    def create_table(self): #creatio le tableau de rapport inserer les titres

        v=0
        list_jaune=[]
        list_bleu=[]
        self.read_data()

        for i in list_names_titles:
            x=list_indice[v]
            v+=1
            self.list_state_titles = ['Nb "G"\n'+x, 'Nb "O"\n'+x, 'Nb "R"\n'+x, 'Nb "NA"\n'+x, 'Nb "NE"\n'+x, '% G\n' + x,
                                  '% O\n' + x,
                                  '% R\n' + x, '% NA\n' + x, '% NE\n' + x, ' ', 'N\n'+x,
                                  'n.\n'+x,
                                  'KPI.1a\n' + x,
                                  'KPI.1b\n' + x, 'Nb"G"\n'+x, 'Nb. "NS_G"\n'+x, 'calcul KPI.1c\n'+x, 'KPI.1c\n' + x,
                                  'Nb "FS"\n'+x, 'Nb "PS"\n'+x, 'Nb "NS"\n'+x,
                                  'calcul KPI.1d\n'+x, 'KPI.1d\n' + x,
                                      ]
            list_jaune.extend([self.list_state_titles[i] for i in [13, 14, 18, 23]])
            list_bleu.extend([self.list_state_titles[i] for i in [5, 6, 7, 8, 9]])

        self.my_red = openpyxl.styles.colors.Color(rgb='00FF0000')
        self.my_green = openpyxl.styles.colors.Color(rgb='0066b615')
        self.my_orange = openpyxl.styles.colors.Color(rgb='00FFA500')
        self.my_gray = openpyxl.styles.colors.Color(rgb='00C0C0C0')
        self.my_bleu = openpyxl.styles.colors.Color(rgb='003399ff')
        self.my_yallow = openpyxl.styles.colors.Color(rgb='00FFFF00')
        my_blue = openpyxl.styles.colors.Color(rgb='001E90FF')
        self.my_black = openpyxl.styles.colors.Color(rgb='00000000')
        self.my_white = openpyxl.styles.colors.Color(rgb='00FFFFFF')
        header_formatfont = Font(bold=True, )
        header_formattxt = Alignment(wrap_text=True)

        workbook = Workbook()
        self.workboo2=workbook
        rapport = workbook.active

        rapport.title = 'Results_eCLEM'
        ref= 'A6:'+str(get_column_letter(len(self.list_report)))+str(len(self.data1['ID'])+6)


        tab = Table(displayName="Table1", ref=ref)
        # I list out the 4 show-xyz options here for reference
        style = TableStyleInfo(
            #name="TableStyleLight18",
            name="TableStyleLight21",
            #name="TableStyleMedium22",

            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
        tab.tableStyleInfo = style
        rapport.add_table(tab)
        x=6
        all_names = []
        for i in list_names_titles:
            for j in i:
                all_names.append(j)
        for i in  range(len(list_report)):
            rapport.cell(x,i+1).value=self.list_report[i]
            rapport.cell(x, i+1).border = thin_border
            if rapport.cell(x,i+1).value in list_head[0] :
                rapport.cell(x,i+1).font = header_formatfont # font gras pour les titres
                rapport.column_dimensions[get_column_letter(i + 1)].width = 15
            if rapport.cell(x,i+1).value in self.list_red:rapport.cell(x,i+1).font = Font(color="FFFF0000")# mette les coulleur rouge
            if rapport.cell(x,i+1).value in list_jaune:
                rapport.cell(x,i+1).fill = PatternFill(patternType='solid', fgColor=self.my_yallow)# mette les coulleur jaune
                rapport.column_dimensions[get_column_letter(i + 1)].width = 15
            if rapport.cell(x,i+1).value in list_bleu:
                rapport.cell(x,i+1).fill = PatternFill(patternType='solid', fgColor=my_blue)# mette les coulleur bleu
                rapport.column_dimensions[get_column_letter(i + 1)].width = 15

            rapport.cell(x-1,i+1).fill = PatternFill(patternType='solid', fgColor=self.my_black)
            rapport.cell(x,i+1).alignment = header_formattxt
            rapport.cell(x, i + 1 ).alignment = Alignment(horizontal='center', vertical='center',
                                                                 wrap_text=True)



            if list_report[i] not in list_head[0] and  list_report[i] not in all_names:
                 rapport.column_dimensions[get_column_letter(i + 1)].width = 15



        self.insert_data(rapport,workbook)




    def red_items(self):
        list_red_index=[]
        wb = openpyxl.load_workbook(filename=self.vfichier1.get())
        sheet_ranges = wb['VehicleScoreCard']
        for i in range(1, len(self.list_original)):
            color_obj = sheet_ranges.cell(row=6, column=i ).font.color
            if color_obj is not None:
                if color_obj.rgb == "FFFF0000": # pour ajouter les titre qui sont on couleur rouge pour les prendre en considiration
                    list_red_index.append(self.list_report.index(self.list_original[i]))

        for i in list_red_index:
            self.list_red .append(self.list_report[i])





    def state2(self, rap2_id, workbook,rap2_kpi_ab,kpi_c,kpi_cv,kpi_d,kpi_dv):#raport2
        self.progress_bar["value"] = 80
        root.update()

        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))

        header_formatfont = Font(bold=True, )
        header_formattxt = Alignment(wrap_text=True)

        rapport = workbook.create_sheet('KPI.1_eCLEM')  # creation le deuxieme rapport


        a = 0



        list_head = [['ID', 'Name', 'Safety', 'MBD', 'SwQA', 'Direction', 'Service','ASWP deployment for'], ]
        list_kpi = ["KPI1.a", "KPI1.b", "KPI1.c", "KPI1.d"]
        list_kpi_dif = ["Deliverable \nCommittment", "Deliverable \nCreation", "SWQA \nCheck-Rate", "SWQA \nStatus"]

        list_indice = ['Project Management', 'Improvement Management', 'Supporting activities Management',
                       'SW Requirements Management',
                       'SW Architecture & Design Management', 'SW Coding/Modeling Management',
                       'SW Verification Management', 'SW Qualification Management',
                       'SW Safety Management', 'Supplier Management', 'GLOBAL']
        y = 0
        yy=0
        z = 0

        for k in range(len(rap2_id[0])):
            for j in range(8):


                for i in range(7):
                    if i == 0:
                        if j<7:

                            c1 = get_column_letter(i + 2 + a) + str(j + 1)
                            c2 = get_column_letter(i + 7 + a) + str(j + 1)
                            c = c1 + ":" + c2

                            rapport.merge_cells(c)
                            rapport.column_dimensions[get_column_letter(i + 1 + a)].width = 40
                            rapport.merge_cells(c)
                        else:
                            c1 = get_column_letter(i + 2 + a) + str(j + 1)
                            c2 = get_column_letter(i + 5 + a) + str(j + 1)
                            c = c1 + ":" + c2
                            rapport.merge_cells(c)
                            rapport.column_dimensions[get_column_letter(i + 1 + a)].width = 40
                            rapport.merge_cells(c)
                            c1 = get_column_letter(i + 6 + a) + str(j + 1)
                            c2 = get_column_letter(i + 7 + a) + str(j + 1)
                            c = c1 + ":" + c2
                            rapport.merge_cells(c)
                            rapport.column_dimensions[get_column_letter(i + 1 + a)].width = 40
                            rapport.merge_cells(c)

                    else:
                        rapport.cell(j + 1, 1+a ).value = list_head[0][j]
                        rapport.cell(1, 2 + a).value = rap2_id[0][y]
                        rapport.cell(2, 2 + a).value = rap2_id[1][y]
                        rapport.cell(3, 2 + a).value = rap2_id[2][y]
                        rapport.cell(4, 2 + a).value = rap2_id[3][y]
                        rapport.cell(5, 2 + a).value = rap2_id[4][y]
                        rapport.cell(6, 2 + a).value = rap2_id[5][y]
                        rapport.cell(j + 10, 1 + a).value = rap2_id[1][y]
                        rapport.cell(j + 10, 1 + 2 + a).border = thin_border

                        rapport.cell(j + 1, 1+ a).border = thin_border

                        if rap2_id[0].index(rap2_id[0][y]) in rap2_id[6] :
                          rapport.cell(7, 2 + a).value = self.list_service[rap2_id[6].index(rap2_id[0].index(rap2_id[0][y]))]

                        rapport.cell(j + 1, i + 1 + a).border = thin_border

                        rapport.column_dimensions[get_column_letter(i + 1 + a)].width = 25
                        #rapport.cell(j + 1,  1 + a).font = Font(color="00FF8C00")
                        #rapport.cell(j + 1,  2 + a).font = Font(color="00FF8C00")
                        rapport.cell(j + 1, 1+a).font = Font(bold=True)


                    rapport.cell(j+1, i + 1 + a).alignment = Alignment(horizontal='center', vertical='center')



            y+=1
            x = 0
            x2=0
            x3=0

            for j in range(11):


                for i in range(7):
                    if j == 0:
                        c1 = get_column_letter(1 + a) + str(8)
                        c2 = get_column_letter(1 + a) + str(9)

                        c = c1 + ":" + c2
                        rapport.merge_cells(c)


                        if i < 4:
                            rapport.cell(j + 9, i + 2 + a).value = list_kpi[i]

                            rapport.cell(j + 9, i + 2 + a).border = thin_border
                            rapport.cell(j + 9, i + 2 + a).alignment = Alignment(horizontal='center',
                                                                                 vertical='center', )

                            rapport.cell(j + 10, i + 2 + a).value = list_kpi_dif[i]
                            rapport.cell(j + 10, i + 2 + a).border = thin_border
                            rapport.cell(j + 10, i + 2 + a).alignment = Alignment(horizontal='center',
                                                                                 vertical='center', )


                            #rapport.cell(j + 8, i+2 + a).font = Font(color="00FF8C00")
                            rapport.cell(j + 10, i+2 + a).font = Font(bold=True,)
                            rapport.cell(j + 10,  1 + a).alignment = Alignment(horizontal='center', vertical='center',)

                            rapport.cell(8,  2 + a).value = "KPI.1 Deliverable \nstatus"
                            rapport.cell(8,  2 + a).font = Font(bold=True,)

                            ##############
                            rapport.cell(8, 6 + a).value = "KPI.2 Process \napplication status"
                            rapport.cell(8, 6 + a).font = Font(bold=True, )

                        if i>4:
                            rapport.cell(j + 9, 7 + a).value = 'Template \napplication '
                            rapport.cell(j + 9, 6 + a).value ='Process \napplication '
                            rapport.cell(j + 9, 7 + a).font = Font(bold=True, )
                            rapport.cell(j + 9, 6 + a).font = Font(bold=True, )
                            rapport.cell(j + 9, 7 + a).border = thin_border
                            rapport.cell(j + 9, 6 + a).border = thin_border
                            rapport.cell(j + 9, 7 + a).alignment = Alignment(horizontal='center', vertical='center', )
                            rapport.cell(j + 9, 6 + a).alignment = Alignment(horizontal='center', vertical='center', )




                    if 0<i<3:
                        if len(rap2_kpi_ab) >0:

                            rapport.cell(j + 11, i + 1 + a).value = rap2_kpi_ab[z][x]
                            rapport.cell(j + 11, i + 1 + a).alignment = Alignment(horizontal='center', vertical='center')
                            try:
                                valeur = float(str(rap2_kpi_ab[z][x]).replace('%', ''))
                                if valeur == 100:
                                    rapport.cell(j + 11, i + 1 + a).fill = PatternFill(patternType='solid',
                                                                                       fgColor=self.my_bleu)
                                if 75 <= valeur < 100:
                                    rapport.cell(j + 11, i + 1 + a).fill = PatternFill(patternType='solid',
                                                                                       fgColor=self.my_green)
                                if 50 <= valeur < 75:
                                    rapport.cell(j + 11, i + 1 + a).fill = PatternFill(patternType='solid',
                                                                                       fgColor=self.my_yallow)
                                if 25 <= valeur < 50:
                                    rapport.cell(j + 11, i + 1 + a).fill = PatternFill(patternType='solid',
                                                                                       fgColor=self.my_orange)
                                if 0 <= valeur < 25:
                                    rapport.cell(j + 11, i + 1 + a).fill = PatternFill(patternType='solid',
                                                                                       fgColor=self.my_red)
                            except ValueError:
                                valeur = str(rap2_kpi_ab[z][x]).replace('%', '')


                                if str(valeur)=='NA':
                                    rapport.cell(j + 11, i + 1 + a).fill = PatternFill(patternType='solid',fgColor=self.my_gray)
                            if str(valeur) == 'NA':
                                rapport.cell(j + 11, i + 1 + a).fill = PatternFill(patternType='solid',
                                                                                   fgColor=self.my_gray)
                                #rapport.column_dimensions[get_column_letter(i + 1 + a)].width = 30

                            x += 1
                    if i==3:
                        if len(kpi_c) >0:
                            if rap2_id[0].index(rap2_id[0][yy]) in kpi_cv[x2]:
                                rapport.cell(j+11, i+1 + a).value = kpi_c[x2][ kpi_cv[x2].index(rap2_id[0].index(rap2_id[0][yy]))]




                                try:
                                    valeur = float(str(kpi_c[x2][ kpi_cv[x2].index(rap2_id[0].index(rap2_id[0][yy]))]).replace('%', ''))

                                    if valeur == 100:
                                        rapport.cell(j + 11, i + 1 + a).fill = PatternFill(patternType='solid',
                                                                                           fgColor=self.my_bleu)
                                    if 75 <= valeur < 100:
                                        rapport.cell(j + 11, i + 1 + a).fill = PatternFill(patternType='solid',
                                                                                           fgColor=self.my_green)
                                    if 50 <= valeur < 75:
                                        rapport.cell(j + 11, i + 1 + a).fill = PatternFill(patternType='solid',
                                                                                           fgColor=self.my_yallow)
                                    if 25 <= valeur < 50:
                                        rapport.cell(j + 11, i + 1 + a).fill = PatternFill(patternType='solid',
                                                                                           fgColor=self.my_orange)
                                    if 0 <= valeur < 25:
                                        rapport.cell(j + 11, i + 1 + a).fill = PatternFill(patternType='solid',
                                                                                           fgColor=self.my_red)
                                    if str(valeur) == 'NA':
                                        rapport.cell(j + 11, i + 1 + a).fill = PatternFill(patternType='solid',
                                                                                           fgColor=self.my_gray)

                                except ValueError:


                                        rapport.cell(j + 11, i + 1 + a).fill = PatternFill(patternType='solid',
                                                                                           fgColor=self.my_gray)

                            rapport.column_dimensions[get_column_letter(i + 1 + a)].width = 30
                            rapport.cell(j + 11, i + 1 + a).alignment = Alignment(horizontal='center',
                                                                                  vertical='center')
                            x2 += 1


                    if i==4:
                        if len(kpi_d) >0:
                            if rap2_id[0].index(rap2_id[0][yy]) in kpi_dv[x3]:
                                f=kpi_d[x3][ kpi_dv[x3].index(rap2_id[0].index(rap2_id[0][yy]))]
                                print(f)

                                if f!='NA' and  f!='NE'  :
                                    f=str(f)+'%'
                                rapport.cell(j+11, i+1 + a).value = f


                                try:
                                    valeur = float(str(kpi_d[x3][kpi_dv[x3].index(rap2_id[0].index(rap2_id[0][yy]))]).replace('%', ''))

                                    if valeur == 100:
                                        rapport.cell(j + 11, i + 1 + a).fill = PatternFill(patternType='solid',
                                                                                           fgColor=self.my_bleu)
                                    if 75 <= valeur < 100:
                                        rapport.cell(j + 11, i + 1 + a).fill = PatternFill(patternType='solid',
                                                                                           fgColor=self.my_green)
                                    if 50 <= valeur < 75:
                                        rapport.cell(j + 11, i + 1 + a).fill = PatternFill(patternType='solid',
                                                                                           fgColor=self.my_yallow)
                                    if 25 <= valeur < 50:
                                        rapport.cell(j + 11, i + 1 + a).fill = PatternFill(patternType='solid',
                                                                                           fgColor=self.my_orange)
                                    if 0 <= valeur < 25:
                                        rapport.cell(j + 11, i + 1 + a).fill = PatternFill(patternType='solid',
                                                                                           fgColor=self.my_red)

                                except ValueError:
                                    if str(kpi_d[x3][kpi_dv[x3].index(rap2_id[0].index(rap2_id[0][yy]))]) == 'NA':
                                        rapport.cell(j + 11, i + 1 + a).fill = PatternFill(patternType='solid',
                                                                                           fgColor=self.my_gray)
                                    else:
                                        rapport.cell(j + 11, i + 1 + a).fill = PatternFill(patternType='solid',
                                                                                           fgColor=self.my_white)

                            rapport.cell(j + 11, i + 1 + a).alignment = Alignment(horizontal='center',
                                                                                  vertical='center')
                            x3+=1
                            rapport.column_dimensions[get_column_letter(i + 1 + a)].width = 20



                    if i>4:
                        c1 = get_column_letter(6 + a) + str(9)
                        c2 = get_column_letter(6 + a) + str(10)

                        c = c1 + ":" + c2
                        rapport.merge_cells(c)
                        c1 = get_column_letter(7 + a) + str(9)
                        c2 = get_column_letter(7 + a) + str(10)

                        c = c1 + ":" + c2
                        rapport.merge_cells(c)



                        rapport.cell(j + 11, i + 1 + a).border = thin_border
                        rapport.cell(j + 11, i + 1 + a).fill = PatternFill(patternType='solid',
                                                                           fgColor=self.my_gray)

                        rapport.column_dimensions[get_column_letter(i + 1 + a)].width = 20



                    rapport.cell(j + 11, 1 + a).value = list_indice[j]
                    rapport.cell(j + 11, i + 1 + a).border = thin_border


                   # rapport.cell(j + 11, i + 1 + a).alignment = Alignment(horizontal='center', vertical='center')
                    #rapport.cell(j+11,1+a).fill = PatternFill(patternType='solid')
                    rapport.cell(j + 11, 1+ a).font = Font(bold=True, )


            a += 8
            z+=1
            yy += 1
        img = openpyxl.drawing.image.Image('fig.png')

        img.width = 600
        img.height = 250
        rapport.add_image(img, 'B24')

        ##########################################################################################################################################################

    def state(self, rapport ):
        self.progress_bar["value"] = 30
        root.update()

        def traiter(args):
            i_itmes = []
            for i in range(len(list_names_titles)):
                i_itmes.append(i)  # pour vérifier est ce que self.valeur est compatible
            w = 6
            a = 0

            self.kpi_raport2 = []
            kpi_instance = []

            for k in range(len(self.list_id_net)):

                w += 1
                s = -1
                kpi_instance = []
                for arg in args:


                    gg = 0
                    oo = 0
                    rr = 0
                    nna = 0
                    nne = 0
                    pper = 0
                    fglobal = '='
                    s += 1
                    arg.sort()

                    ffs = 0
                    pps = 0
                    nns = 0
                    nns_g = 0
                    NG=0

                    for i, j in zip(i_itmes, arg):
                        if i in self.valeur:
                            g = 0
                            o = 0
                            r = 0
                            na = 0
                            ne = 0
                            per = 0
                            fs=0
                            ps=0
                            ns=0
                            ns_g=0
                            N=0


                            for n in self.list_red:
                                if n in list_names_titles[i]:
                                    if n in list_id_FullSW:

                                        index=self.list_id_score.index(self.list_id_net[k])
                                        if self.list_SwQA_score[index].upper()=='FULL SW' and  self.list_mbd_score[index].upper() =='YES':



                                            N+=list_yes_FullSW[list_id_FullSW.index(n)]
                                            NG+=list_yes_FullSW[list_id_FullSW.index(n)]

                                        if self.list_SwQA_score[index].upper()=='FULL SW' and  self.list_mbd_score[index].upper() =='NO':
                                             N+=list_no_FullSW[list_id_FullSW.index(n)]
                                             NG += list_no_FullSW[list_id_FullSW.index(n)]
                                        if self.list_SwQA_score[index].upper()=='PA':
                                             N+=list_pa_FullSW[list_id_FullSW.index(n)]
                                             NG += list_pa_FullSW[list_id_FullSW.index(n)]


                                    if self.list_id_net[k] in list(self.data3['ID CLEM O52_ASWP']):
                                        f=self.data3[n][list(self.data3['ID CLEM O52_ASWP']).index(self.list_id_net[k])]
                                        if str(f).upper()=='FS':
                                            fs+=1
                                            ffs+=1
                                        if str(f).upper()=='PS':
                                            ps+=1
                                            pps+=1

                                        if str(f).upper()=='NS':
                                            ns+=1
                                            nns+=1
                                        if str(f).upper()=='NS_G':
                                            ns_g+=1
                                            nns_g+=1

                                    per += 1
                                    pper += 1
                                    instance = list(self.data1[n])[k]

                                    if str(instance).strip() == 'G':
                                        g += 1
                                        gg += 1
                                        rapport.cell(w, self.list_report.index(n) + 1).fill = PatternFill(
                                            patternType='solid', fgColor=self.my_green)

                                    if str(instance).strip() == 'O':
                                        o += 1
                                        oo += 1
                                        rapport.cell(w, self.list_report.index(n) + 1).fill = PatternFill(
                                            patternType='solid', fgColor=self.my_orange)
                                    if str(instance).strip() == 'R':
                                        r += 1
                                        rr += 1

                                        rapport.cell(w, self.list_report.index(n) + 1).fill = PatternFill(
                                            patternType='solid', fgColor=self.my_red)
                                    if str(instance).strip() == '':

                                        rapport.cell(w, self.list_report.index(n) + 1).value='NO DATA'



                                    if str(instance).strip() == 'NA':
                                        na += 1
                                        nna += 1
                                        rapport.cell(w, self.list_report.index(n) + 1).value='NA'
                                        rapport.cell(w, self.list_report.index(n) + 1).fill = PatternFill(
                                            patternType='solid', fgColor=self.my_gray)

                                    if str(instance).strip() == 'NE':
                                        rapport.cell(w, self.list_report.index(n) + 1).fill = PatternFill(
                                            patternType='solid', fgColor=self.my_white)
                                        ne += 1
                                        nne += 1

                            ################################lie a tout les pertie qui des titre rouge###############################
                            go = [g, o, r, na, ne]
                            fo=[fs,ps,ns,ns_g]
                            if sum(go)==0:
                                rapport.cell(w, j + 1).value = 'NO DATA'
                                rapport.cell(w, j + 6).value = 'NO DATA'
                                rapport.cell(w, j + 7).value = ''
                                rapport.cell(w, j + 8).value = 'NO DATA'  # N
                                rapport.cell(w, j + 9).value = 'NO DATA'  # n
                                rapport.cell(w, j + 10).value = 'NO DATA'  # n
                                rapport.cell(w, j + 11).value = 'NO DATA'  # n

                                kpi_instance.append('NO DATA')
                                rapport.cell(w, j + 12).value = 'NO DATA' # n-Nb"G"
                                rapport.cell(w, j + 13).value = 'NO DATA' # n_NB"NS_G"
                                rapport.cell(w, j + 14).value = 'NO DATA'
                                rapport.cell(w, j + 16).value = 'NO DATA'  # FS########
                                rapport.cell(w, j + 17).value = 'NO DATA'  # FS########
                                rapport.cell(w, j + 18).value = 'NO DATA'  # FS########
                                rapport.cell(w, j + 19).value = 'NO DATA'  # FS########
                                if s == len(go) - 1:
                                    if a < len(list_indice):
                                        rapport.cell(4, j + 15).value = list_indice[
                                            a]  # mette le titre des partie des tableau  a partir de list list_indice
                                        rapport.cell(4, j + 15).font = Font(bold=True, size=18)
                                        rapport.cell(4, j - 3).value = list_indice[
                                            a]  # mette le titre des partie des tableau  a partir de list list_indice
                                        rapport.cell(4, j - 3).font = Font(bold=True, size=18)
                                        a += 1

                            else:
                                    rapport.cell(w, j + 1).value = go[s]

                                    fx1 = '=NB.SI(' + str(get_column_letter(j - 5)) + str(w) + ':' + str(
                                        get_column_letter(j)) + str(w) + ',"O")'
                                    # fx2='=LEN( '+str(get_column_letter(j -5))+str(w)+')-LEN(SUBSTITUTE('+ str(get_column_letter(j -5))+str(w)+',"z",""))'
                                    # rapport.cell(w, j + 1).value = fx2

                                    # rapport.cell(w, j + 1).value = '= NB.SI.ENS(I7:N7;"O")'

                                    # rapport.cell(w, j + 6).value = str(round(100*go[s]/per,2))+"%"
                                    fx6 = '=' + str(get_column_letter(j + 1)) + str(w) + '/' + str(per)

                                    # rapport.cell(w, j + 6).value = fx6
                                    rapport.cell(w, j + 6).value = str(round((100 * go[s] / per), 2)) + "%"
                                    rapport.cell(w, j + 1).border = thin_border
                                    rapport.cell(w, j + 6).border = thin_border

                                    if s == len(go) - 1:
                                        if a < len(list_indice):
                                            rapport.cell(4, j + 15).value = list_indice[a]  # mette le titre des partie des tableau  a partir de list list_indice
                                            rapport.cell(4, j + 15).font = Font(bold=True,size=18 )
                                            rapport.cell(4, j - 3).value = list_indice[a]  # mette le titre des partie des tableau  a partir de list list_indice
                                            rapport.cell(4, j -3).font = Font(bold=True, size=18)
                                            a += 1

                                        rapport.cell(w, j + 8).border = thin_border
                                        rapport.cell(w, j + 9).border = thin_border
                                        rapport.cell(w - 6, j + 7).fill = PatternFill(patternType='solid',
                                                                                      fgColor=self.my_gray)
                                        rapport.cell(w, j + 7).fill = PatternFill(patternType='solid', fgColor=self.my_gray)
                                        rapport.cell(w - 6, j + 21).fill = PatternFill(patternType='solid',
                                                                                       fgColor=self.my_black)  # le vide noir entre les changement
                                        rapport.cell(w, j + 21).fill = PatternFill(patternType='solid',
                                                                                   fgColor=self.my_black)  # le vide noir entre les changement
                                        rapport.cell(w, j + 8).value = N # N
                                        rapport.cell(w, j + 9).value = N # n
                                        if N>0:
                                            rapport.cell(w, j + 10).value = str(per / per * 100) + '%'  # KPI.1a Project Management###########################################################

                                            rapport.cell(w, j + 11).value = str(round(((go[0] * 1 + go[1] * 0.5 + go[2] * 0) / per * 100), 2)) + '%'  # KPI.1bProject Management

                                            kpi_instance.append(str(per / per * 100) + '%')
                                            kpi_instance.append(str(round(((go[0] * 1 + go[1] * 0.5 + go[2] * 0) / per * 100), 2)) + '%')
                                        else:
                                            rapport.cell(w, j + 10).value ='NA'

                                            rapport.cell(w, j + 11).value ='NA'
                                            kpi_instance.append('NA')
                                            kpi_instance.append('NA')



                                        #rapport.cell(w, j + 12).value = sum(go) - go[2]  # n-Nb"R"
                                        rapport.cell(w, j + 12).value =go[0]  # n-Nb"G"
                                        rapport.cell(w, j + 13).value = fo[3] #  n_NB"NS_G"

                                        try:
                                            rapport.cell(w, j + 14).value = str(
                                                round((go[0]-fo[3])*100/go[0] )) + '%' #calcul KPI.1c
                                        except ZeroDivisionError as error:
                                            rapport.cell(w, j + 14).value ='0.00%'




                                        rapport.cell(w, j + 16).value = fo[0] #FS########
                                        rapport.cell(w, j + 17).value = fo[1]
                                        rapport.cell(w, j + 18).value = fo[2]
                                        # fonction pour automatiquer de calcule


                                        fx19 = '(' + str(get_column_letter(j + 16)) + str(w) + '*1+' + str(
                                            get_column_letter(j + 17)) + str(w) + '*0.5+' + str(
                                            get_column_letter(j + 18)) + str(w) + '*0)*100/' +"("+ str(
                                            get_column_letter(j + 12)) + str(w) +'-'+ str(get_column_letter(j + 13)) + str(w)+")" # pour  automatiser

                                        fxglo = '(' + str(get_column_letter(j + 16)) + str(w) + '*1+' + str(
                                            get_column_letter(j + 17)) + str(w) + '*0.5+' + str(
                                            get_column_letter(j + 18)) + str(w) + '*0)*100/' + str(
                                            get_column_letter(j + 13)) + str(w) + '+'  # pour  automatiser
                                        fglobal += fxglo


                                        if go[0]-fo[3]>0:
                                            rapport.cell(w,j + 19).value = '=IF(' + fx19 + ' > 0," "' + '& ' + fx19 + ' & "%",' + fx19 + ' & "%")' # calcule KPI.1D

                                        else:
                                            rapport.cell(w, j + 19).value ='0.00%'







                        ########################################partie global lie a mist gooo############################################


                        elif i == i_itmes[-1]:  # i est l index de  dernier parier de GLOBAL
                            goo = [gg, oo, rr, nna, nne]
                            foo = [ffs, pps, nns, nns_g]
                            if sum(goo) == 0:
                                rapport.cell(w, j + 1).value = 'NO DATA'
                                rapport.cell(w, j + 6).value = 'NO DATA'
                                rapport.cell(w, j + 8).value = 'NO DATA'  # N
                                rapport.cell(w, j + 9).value = 'NO DATA'  # n
                                rapport.cell(w, j + 10).value = 'NO DATA'  # n
                                rapport.cell(w, j + 11).value = 'NO DATA'  # n

                                kpi_instance.append('NO DATA')
                                rapport.cell(w, j + 12).value = 'NO DATA'  # n-Nb"G"
                                rapport.cell(w, j + 13).value = 'NO DATA'  # n_NB"NS_G"
                                rapport.cell(w, j + 14).value = 'NO DATA'
                                rapport.cell(w, j + 16).value = 'NO DATA'  # FS########
                                rapport.cell(w, j + 17).value = 'NO DATA'  # FS########
                                rapport.cell(w, j + 18).value = 'NO DATA'  # FS########
                                rapport.cell(w, j + 19).value = 'NO DATA'  # FS########
                                if s == len(goo) - 1:

                                    if a < len(list_indice):
                                        rapport.cell(4, j + 15).value = list_indice[
                                            a]  # mette le titre des partie des tableau  a partir de list list_indice
                                        rapport.cell(4, j + 15).font = Font(bold=True, size=18)
                                        rapport.cell(4, j - 3).value = list_indice[
                                            a]  # mette le titre des partie des tableau  a partir de list list_indice
                                        rapport.cell(4, j - 3).font = Font(bold=True, size=18)

                                        a += 1
                            else:
                                    rapport.cell(w, j + 1).value = goo[s]
                                    fx6 = '=' + str(get_column_letter(j + 1)) + str(w) + '/' + str(pper) + '*100 '

                                    #rapport.cell(w, j + 6).value = fx6
                                    rapport.cell(w, j + 6).value = str(round((100 * goo[s] / pper), 2)) + "%"
                                    rapport.cell(w, j + 1).border = thin_border
                                    rapport.cell(w, j + 6).border = thin_border
                                    if s == len(goo) - 1:

                                        if a < len(list_indice):
                                            rapport.cell(4, j + 15).value = list_indice[
                                                a]  # mette le titre des partie des tableau  a partir de list list_indice
                                            rapport.cell(4, j + 15).font = Font(bold=True, size=18)
                                            rapport.cell(4, j - 3).value = list_indice[
                                                a]  # mette le titre des partie des tableau  a partir de list list_indice
                                            rapport.cell(4, j - 3).font = Font(bold=True, size=18)

                                            a += 1

                                        rapport.cell(w, j + 8).border = thin_border
                                        rapport.cell(w, j + 9).border = thin_border
                                        rapport.cell(w - 6, j + 7).fill = PatternFill(patternType='solid',
                                                                                      fgColor=self.my_gray)
                                        rapport.cell(w, j + 7).fill = PatternFill(patternType='solid', fgColor=self.my_gray)
                                        rapport.cell(w - 6, j + 21).fill = PatternFill(patternType='solid',
                                                                                       fgColor=self.my_black)  # le vide noir entre les changement
                                        rapport.cell(w, j + 21).fill = PatternFill(patternType='solid',
                                                                                   fgColor=self.my_black)  # le vide noir entre les changement
                                        rapport.cell(w, j + 8).value = NG  # N
                                        rapport.cell(w, j + 9).value = NG  # n
                                        if NG> 0:
                                            rapport.cell(w, j + 10).value = str(
                                                pper / pper * 100) + '%'  # KPI.1a Project Management
                                            rapport.cell(w, j + 11).value = str(
                                                round(((goo[0] * 1 + goo[1] * 0.5 + goo[2] * 0) / pper * 100),
                                                      2)) + '%'  # KPI.1bProject Management
                                            kpi_instance.append(str(pper / pper * 100) + '%')
                                            kpi_instance.append(str(
                                                round(((goo[0] * 1 + goo[1] * 0.5 + goo[2] * 0) / pper * 100),
                                                      2)) + '%')
                                        else:
                                            rapport.cell(w, j + 10).value = 'NA'

                                            rapport.cell(w, j + 11).value = 'NA'
                                            kpi_instance.append('NA')
                                            kpi_instance.append('NA')





                                        rapport.cell(w, j + 12).value = goo[0]  # n-Nb"G"
                                        rapport.cell(w, j + 13).value = foo[3]  # n_NB"NS_G"
                                        try:
                                            rapport.cell(w, j + 14).value = str(
                                                round((goo[0]-foo[3])*100/goo[0] )) + '%' #calcul KPI.1c
                                        except ZeroDivisionError as error:
                                            rapport.cell(w, j + 14).value ='0.00%'

                                        rapport.cell(w, j + 16).value = foo[0]  # FS########
                                        rapport.cell(w, j + 17).value = foo[1]
                                        rapport.cell(w, j + 18).value = foo[2]
                                        # fonction pour automatiquer de calcule


                                        fxglo2 = '(' + str(get_column_letter(j + 16)) + str(w) + '*1+' + str(
                                            get_column_letter(j + 17)) + str(w) + '*0.5+' + str(
                                            get_column_letter(j + 18)) + str(w) + '*0)*100/' +"(" +str(
                                            get_column_letter(j + 12)) + str(
                                            w)+'-'+str(
                                            get_column_letter(j + 13)) + str(
                                            w)+")"
                                        if goo[0]-foo[3] >0:
                                            rapport.cell(w, j + 19).value = '=IF(' + fxglo2 + ' > 0," "' + '& ' + fxglo2 + ' & "%",' + fxglo2 + ' & "%")' # calcule KPI.1D
                                        else:
                                            rapport.cell(w, j + 19).value ='0.00%' # KPI.1d CLACULE




                                # rapport.cell(w, j + 19).value = 'IF(+'+str(fxglo2)+'  > 0,'+""+str(fxglo2)+"'%'"+',"False")'



                        ###############################################parite lie a les elemnt qui ont aps aucun titre rouge
                        else:

                            go = [1, 2, 3, 4, 5]
                            rapport.cell(w, j + 1).value = 0
                            rapport.cell(w, j + 6).value = "0%"
                            rapport.cell(w, j + 1).border = thin_border
                            rapport.cell(w, j + 6).border = thin_border
                            if s == len(go) - 1:
                                if a < len(list_indice):
                                    rapport.cell(4, j + 15).value = list_indice[
                                        a]  # mette le titre des partie des tableau  a partir de list list_indice

                                    rapport.cell(4, j + 15).font = Font(bold=True, size=18)
                                    rapport.cell(4, j - 3).value = list_indice[
                                        a]  # mette le titre des partie des tableau  a partir de list list_indice
                                    rapport.cell(4, j - 3).font = Font(bold=True, size=18)
                                    a += 1
                                rapport.cell(w, j + 8).border = thin_border
                                rapport.cell(w, j + 9).border = thin_border
                                rapport.cell(w - 6, j + 7).fill = PatternFill(patternType='solid',
                                                                              fgColor=self.my_gray)  # c'est pour siparation en gray '
                                rapport.cell(w, j + 7).fill = PatternFill(patternType='solid', fgColor=self.my_gray)
                                rapport.cell(w - 6, j + 21).fill = PatternFill(patternType='solid',
                                                                               fgColor=self.my_black)
                                rapport.cell(w, j + 21).fill = PatternFill(patternType='solid',
                                                                           fgColor=self.my_black)
                                rapport.cell(w, j + 8).value = 0  # N
                                rapport.cell(w, j + 9).value = 0  # n
                                rapport.cell(w, j + 10).value = 'NA'  # KPI.1a Project Management
                                rapport.cell(w, j + 11).value = 'NA'
                                kpi_instance.append('NA')# pour l'envoyer
                                kpi_instance.append('NA')

                                rapport.cell(w, j + 12).value = 0  # n-Nb"R"
                                rapport.cell(w, j + 13).value = 0  # V
                                if go[2] != 4:
                                    rapport.cell(w, j + 14).value = '0%'
                                else:
                                    rapport.cell(w, j + 14).value = '0' + '%'



                self.kpi_raport2.append(kpi_instance)



        v=0
        n=[]
        g=[]
        o=[]
        r=[]
        na=[]

        for i in list_names_titles:
            x=list_indice[v]
            v+=1


            g.append('Nb "G"\n'+x)
            o.append('Nb "O"\n'+x)
            r.append('Nb "R"\n'+x)
            na.append( 'Nb "NA"\n'+x)
            n.append('Nb "NE"\n' + x)
        n_index=[]
        g_index=[]
        o_index=[]
        r_index=[]
        na_index=[]
        for k in g :
            g_index.append(self.list_report.index(k))
        for k in n :
            n_index.append(self.list_report.index(k))

        for k in o :
            o_index.append(self.list_report.index(k))
        for k in r :
            r_index.append(self.list_report.index(k))
        for k in na :
            na_index.append(self.list_report.index(k))

        list_c = [g_index, o_index, r_index, na_index, n_index]

        self.valeur = []
        self.list_c2 = []  # list des index des coloone de state des couleur

        for i in self.list_report:
            for j in list_names_titles:
                if i in j and i in self.list_red:
                    if list_names_titles.index(j) not in self.valeur:
                        self.valeur.append(list_names_titles.index(j))
                    for p in list_c:
                        self.z = []
                        if p in list_c:
                            for h in range(len(p)):
                                if p[h] not in self.z:
                                    self.z.append(p[h])
                            if self.z not in self.list_c2:
                                self.list_c2.append(self.z)
                        else:
                            for h in range(len(p)):
                                if h % 2 == 0:
                                    if p[h] not in self.z:
                                        self.z.append(p[h])
                            if self.z not in self.list_c2:
                                self.list_c2.append(self.z)

                        # for h in range(len(r_index)) :
                        # if h %2==0:
                        # if r_index[h] not in self.f:
                        #  self.f.append(r_index[h])

        w = 6
        traiter(list_c)



if __name__== "__main__":
    root = Tk()


    report(root)
    root.mainloop()
