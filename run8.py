from tkinter import *
from tkinter import messagebox
from tkinter.filedialog import askopenfilename,askdirectory
from tkinter import messagebox

import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, colors

from openpyxl.styles import Font, Alignment, PatternFill, colors ,Color
from openpyxl.worksheet.table import Table, TableStyleInfo

from openpyxl.styles.borders import Border, Side
from openpyxl.utils import get_column_letter
from tkinter.ttk import Progressbar
from pathlib import Path

from openpyxl.utils import FORMULAE

import copy
thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

list_id_FullSW=['MAN.I_01', 'MAN.P_01', 'MAN.P_01a', 'MAN.P_02', 'MAN.P_03', 'MAN.P_04', 'MAN.P_05', 'MAN.P_06', 'MAN.S_01',
                             'MAN.S_02', 'MAN.S_03', 'MAN.S_04', 'MAN.S_05', 'MAN.S_06', 'MAN.S_07', 'MAN.S_08', 'MAN.S_09', 'SAF_01', 'SAF_02',
                             'SAF_03', 'SUP_01', 'SUP_01a', 'SUP_02', 'SUP_04', 'SUP_05', 'SUP_06', 'SUP_07', 'SUP_08', 'SUP_09', 'SUP_10', 'SUP_10a',
                             'SUP_11', 'SWE.1_01', 'SWE.1_02', 'SWE.1_02a', 'SWE.1_03', 'SWE.2_01', 'SWE.2_02', 'SWE.3_01', 'SWE.3_02', 'SWE.4_01', 'SWE.4_02', 'SWE.4_03',
                             'SWE.4_10', 'SWE.6_01', 'SWE.6_02', 'SWE.6_03', 'SWE.6_04', 'SWE.6_05', 'SWE.6_06', 'SWE_01', 'SWE_02', 'SWE_03', 'SWE_04']

list_yes_FullSW=[1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1]
list_no_FullSW=[1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 0, 0, 0, 1, 1, 1, 1, 1, 0, 0, 1, 1, 0, 0, 1, 1, 0, 0, 1, 0, 0, 0, 0, 0, 0, 0, 0, 1, 0, 1, 0, 1, 1, 1, 0, 0, 1]
list_pa_FullSW=[0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 1, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 1, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 1, 0, 1, 0, 1, 1, 0, 0, 0, 0]


list_indice=['Project Management','Improvement Management','Supporting activities Management','SW Requirements Management',
             'SW Architecture & Design Management','SW Coding/Modeling Management','SW Verification Management','SW Qualification Management',
             'SW Safety Management','Supplier Management','GLOBAL']

#les titrs des paritie de rapport final



list_indice3=['ACQ.3 Contract Agreement','ACQ.4 Supplier Monitoring','All SWE.X','MAN.3 Project Management','MAN.5 Risk Management','MAN.6 Measurement',
   'PIM.3 Process Improvement','Specific Safety','SPL.2 Product Release','SUP.1 Quality Assurance','SUP.10 Change Request Management','SUP.8 Configuration Management',
   'SUP.9 Problem Resolution Management','SWE.1 Software Requirements Analysis','SWE.2 Software Architectural Design','SWE.3 Software Detailed Design and Unit Construction',
   'SWE.4 Software Unit Verification','SWE.6 Software Qualification Test','GLOBAL']




list_names_titles3=[['MAN.S_01', 'MAN.S_02', 'MAN.S_03', 'MAN.S_04', 'MAN.S_05', 'MAN.S_09'], ['MAN.S_06', 'MAN.S_07', 'MAN.S_08'],
                    ['SWE_01', 'SWE_04'], ['MAN.P_01', 'MAN.P_01a', 'MAN.P_02', 'MAN.P_03'], ['MAN.P_04', 'MAN.P_05'], ['MAN.P_06'],
                    ['MAN.I_01'], ['SAF_01', 'SAF_02', 'SAF_03'], ['SWE_02', 'SWE_03'], ['SUP_01', 'SUP_01a', 'SUP_02'], ['SUP_09', 'SUP_10', 'SUP_10a', 'SUP_11'],
                    ['SUP_04', 'SUP_05', 'SUP_06', 'SUP_07'], ['SUP_08'], ['SWE.1_01', 'SWE.1_02', 'SWE.1_02a', 'SWE.1_03'],
                    ['SWE.2_01', 'SWE.2_02'], ['SWE.3_01', 'SWE.3_02'], ['SWE.4_01', 'SWE.4_02', 'SWE.4_03', 'SWE.4_10'],
                    ['SWE.6_01', 'SWE.6_02', 'SWE.6_03', 'SWE.6_04', 'SWE.6_05', 'SWE.6_06'], []]

list_indec=[]

list_tit=[]

list_state_titles2 = ['Nb "G"', 'Nb "O"', 'Nb "R"', 'Nb "NA"', 'Nb "NE"', '% G\nProject Management',
                     '% O\nProject Management',
                     '% R\nProject Management', '% NA\nProject Management', '% NE\nProject Management', ' ', 'N',
                     'n',
                     'KPI.1a\nProject Management',
                     'KPI.1b\nProject Management', 'n-Nb"R"', 'Nb "G"', 'Unnamed', 'KPI.1c\nProject Management',
                     'Nb "G"', 'Nb "O"', 'Nb "R"','Nb "NA"', 'Nb "NE"',
                     'Unnamed', 'KPI.1d\nProject Management',
                     ]  # length=27 # est ajouter pour regler le dicalage par ce que la fonctio inumirate touve un problem si il ' a
                         # 'Nb "R"', 'Nb "NA"', 'Nb "NE" deux fois on l'ajoute pour regler le dicalage dans liste


list_head = [['ID', 'Name', 'Safety', 'MBD', 'SwQA', 'Direction', 'Service'], ] # length=7 les titres de présentaion de rapport
list_names_titles = [
    ['MAN.P_01a', 'MAN.P_01', 'MAN.P_02', 'MAN.P_03', 'MAN.P_04', 'MAN.P_05', 'MAN.P_06'],
    ['MAN.I_01', ],
    ['SUP_01a', 'SUP_01', 'SUP_02', 'SUP_04', 'SUP_05', 'SUP_06', 'SUP_07', 'SUP_08', 'SUP_09', 'SUP_10a', 'SUP_10',
     'SUP_11', 'SWE_01', 'SWE_04', ],
    ['SWE.1_01', 'SWE.1_02a', 'SWE.1_02', 'SWE.1_03', ],
    ['SWE.2_01', 'SWE.2_02', ],
    ['SWE.3_01', 'SWE.3_02', ],
    ['SWE.4_01', 'SWE.4_02', 'SWE.4_03', 'SWE.4_10', ],
    ['SWE.6_01', 'SWE.6_02', 'SWE.6_03', 'SWE.6_04', 'SWE.6_05', 'SWE.6_06', 'SWE_02', 'SWE_03', ],
    ['SAF_01', 'SAF_02', 'SAF_03', ],
    ['MAN.S_01', 'MAN.S_02', 'MAN.S_03', 'MAN.S_04', 'MAN.S_05', 'MAN.S_06', 'MAN.S_07', 'MAN.S_08',
     'MAN.S_09', ],[] ]# length=10 groupe of names names





def create_all_list():


    a = []
    v=0
    for i in list_names_titles:# creer self.list_report
        x=list_indice[v]
        list_state_titles = ['Nb "G"\n'+x, 'Nb "O"\n'+x, 'Nb "R"\n'+x, 'Nb "NA"\n'+x, 'Nb "NE"\n'+x, '% G\n' + x,
                                  '% O\n' + x,
                                  '% R\n' + x, '% NA\n' + x, '% NE\n' + x, ' ', 'N\n'+x,
                                  'n.\n'+x,
                                  'KPI.1a\n' + x,
                                  'KPI.1b\n' + x, 'Nb"G"\n'+x, 'Nb. "NS_G"\n'+x, 'calcul KPI.1c\n'+x, 'KPI.1c\n' + x,
                                  'Nb "FS"\n'+x, 'Nb "PS"\n'+x, 'Nb "NS"\n'+x,
                                  'calcul KPI.1d\n'+x, 'KPI.1d\n' + x,
                                  ]
        v+=1
        instance = []
        instance = i + list_state_titles
        for j in instance:
            a.append(j)

        a.append(" ")
    a=list_head[0]+a
    b = []
    for i in list_names_titles:# creer la list self.list_original
        for j in i:
            b.append(j)
        b.append(" ")
    b=list_head[0]+b
    c=[]

    for i in list_names_titles:#creer self.list_report2 qui va regler le dicalage dans la liste
        instance = []
        instance = i + list_state_titles2
        for j in instance:
            c.append(j)

        c.append(" ")
    c = list_head[0] + c

    d = []
    v = 0
    for i in list_names_titles3:  # creer self.list_report
        x = list_indice3[v]
        list_state_titles = ['Nb "G"\n'+x, 'Nb "O"\n'+x, 'Nb "R"\n'+x, 'Nb "NA"\n'+x, 'Nb "NE"\n'+x, '% G\n' + x,
                                  '% O\n' + x,
                                  '% R\n' + x, '% NA\n' + x, '% NE\n' + x, ' ', 'N\n'+x,
                                  'n.\n'+x,
                                  'KPI.1a\n' + x,
                                  'KPI.1b\n' + x, 'n-Nb"G"\n'+x, 'Nb. "NS_G"\n'+x, 'calcul KPI.1c\n'+x, 'KPI.1c\n' + x,
                                  'Nb "FS"\n'+x, 'Nb "PS"\n'+x, 'Nb "NS"\n'+x,
                                  'calcul KPI.1d\n'+x, 'KPI.1d\n' + x,
                             ]
        v += 1
        instance = []
        instance = i + list_state_titles
        for j in instance:
            d.append(j)

        d.append(" ")
    d = list_head[0] + d

    return a, b ,c ,d






list_original=create_all_list()[1] # c'est la list des titres de tableau de fichier score....
list_report=create_all_list()[0] # c'est le global des titrs qui se trouve dans  le rapport final
list_report2=create_all_list()[2] # c'est pour regler le dicalage ici ona utiliser list_state_titles2 unitile
 # list_red c'est l'esemble des noms qui ont couleur rouge
list_report3=create_all_list()[3]


class report3:



    def __init__(self,data1,data2,data3,workbook,list_red ):
        self.list_red=list_red

        self.workbook=workbook
        self.data1=data1
        self.data2=data2
        self.data3=data3
        self.list_original=create_all_list()[1] # c'est la list des titres de tableau de fichier score....

        self.list_report2=create_all_list()[2] # c'est pour regler le dicalage ici ona utiliser list_state_titles2 unitile
        # list_red c'est l'esemble des noms qui ont couleur rouge
        self.list_report3 = create_all_list()[3]
        self.create_table()



    # lire les donnéés avec pandas



    def read_data(self):


        self.list_service = list(self.data2['Service'])
        self.list_id = list(self.data2['ID CLEM O52_ASWP'])
        self.list_KPI1_c = list(self.data2['KPI1.c'])
        self.list_KPI1_d = list(self.data2['KPI1.d'])

        self.list_id_net=[]
        self.list_plus=[]
        for i in list(self.data1['ID']):
            if i !='nan':
                self.list_id_net.append(i)




    def insert_KPI(self, rapport3, workbook): #inserer les données de KPI qui se trouvr dans le fichiers suivi.. sheet ..stratigie  dans la fonction reand data


        header_formattxt = Alignment(wrap_text=True)

        list_indice1=['KPI.1c\n'+i for i in list_indice3]
        list_indice2=['KPI.1d\n'+i for i in list_indice3]

        self.kpi_c=[]
        self.kpi_cv=[] #pour rapport3
        self.kpi_d=[]
        self.kpi_dv=[]
        kpic_instance=[]
        for i in range(len(self.list_report3)):# l'objectif c'est trouver lindex de KPI dans le tableau de rapport pour
            # inserer les donné dans la coloune qui a ce index
            x = 6
            kpic_instance = []

            kpicv_instance = []

            kpid_instance = []

            kpidv_instance = []

            if self.list_report3[i] in list_indice1:


                for m, n in zip(self.list_KPI1_c, self.list_id):


                    x += 1
                    if n in list(self.data1['ID']):
                        v = list(self.data1['ID']).index(int(n)) + 7
                        k = "Only check rules eCLEM fill"
                        if k in str(m):

                            rapport3.cell(v, i + 1).value = "Only check \n" \
                                                           "rules eCLEM \nfill"
                            rapport3.cell(v, i + 1).border = thin_border
                            rapport3.cell(v, i + 1).alignment = header_formattxt
                            rapport3.cell(v, i + 1).fill = PatternFill(patternType='solid', fgColor=self.my_gray)
                            kpic_instance.append("Only check \n" \
                                                 "rules eCLEM \nfill")

                            kpicv_instance.append(v-7)

                        else:

                            # rapport.cell(v, i + 1).value = str(rapport.cell(v, i).value)
                            if str(rapport3.cell(v, i).value) == '0.00%' or str(rapport3.cell(v, i).value) == '0%':
                                rapport3.cell(v, i + 1).value = 'NA'
                                rapport3.cell(v, i + 1).fill = PatternFill(patternType='solid', fgColor=self.my_gray)
                                kpic_instance.append('NA')

                            else:
                                rapport3.cell(v, i + 1).value = str(rapport3.cell(v, i).value)
                                rapport3.cell(v, i + 1).fill = PatternFill(patternType='solid', fgColor=self.my_yallow)
                                kpic_instance.append(str(rapport3.cell(v, i).value))
                            rapport3.cell(v, i + 1).border = thin_border


                            kpicv_instance.append(v - 7)



                        rapport3.column_dimensions[get_column_letter(i + 1)].width = 15 # ici la dimmention des case de KPI


                self.kpi_c.append(kpic_instance)
                self.kpi_cv.append(kpicv_instance)





            if self.list_report3[i] in list_indice2:
                for m, n in zip(self.list_KPI1_d, self.list_id):
                    x += 1
                    if n in list(self.data1['ID']):
                        v = list(self.data1['ID']).index(int(n)) + 7
                        k = "NA"
                        try:
                            b = int(rapport3.cell(v, i - 1).value) + int(rapport3.cell(v, i - 2).value) + int(
                                rapport3.cell(v, i - 3).value)
                        except TypeError:
                            b = 1
                        if k in str(m) and b!=0:

                            rapport3.cell(v, i + 1).value = 'NA'
                            rapport3.cell(v, i + 1).border = thin_border
                            rapport3.cell(v, i + 1).alignment = header_formattxt
                            rapport3.cell(v, i + 1).fill = PatternFill(patternType='solid', fgColor=self.my_gray)
                            kpidv_instance.append(v - 7)
                            kpid_instance.append('NA')

                        else:
                            if (int(rapport3.cell(v, i-7).value) - int(rapport3.cell(v, i-6).value))!=0:
                                a=int((int(rapport3.cell(v, i-3).value) * 1 + int(rapport3.cell(v, i-2).value )* 0.5 + int(rapport3.cell(v, i-1).value )* 0) * 100 / (int(rapport3.cell(v, i-7).value) - int(rapport3.cell(v, i-6).value)))
                                b = int(rapport3.cell(v, i - 1).value) + int(rapport3.cell(v, i - 2).value) + int(
                                    rapport3.cell(v, i - 3).value)
                                if a==0 and b!=0:

                                    rapport3.cell(v, i + 1).value = 'NA'
                                    rapport3.cell(v, i + 1).fill = PatternFill(patternType='solid', fgColor=self.my_gray)
                                    kpid_instance.append('NA')

                                elif b == 0 and str(rapport3.cell(v, i - 4).value)!="Only check \nrules eCLEM \nfill" and str(rapport3.cell(v, i - 4).value)!='NA' :
                                    rapport3.cell(v, i + 1).value = 'NE'
                                    rapport3.cell(v, i + 1).fill = PatternFill(patternType='solid', fgColor=self.my_white)
                                    kpid_instance.append('NE')

                                elif str(rapport3.cell(v, i - 4).value)=="Only check \nrules eCLEM \nfill" or str(rapport3.cell(v, i - 4).value)=='NA':
                                    rapport3.cell(v, i + 1).value = 'NA'
                                    rapport3.cell(v, i + 1).fill = PatternFill(patternType='solid',
                                                                              fgColor=self.my_gray)
                                    kpid_instance.append('NA')
                                elif b != 0 and str(rapport3.cell(v, i - 4).value).strip() != "Only check \nrules eCLEM \nfill" and str(rapport3.cell(v, i - 4).value) != 'NA':

                                    rapport3.cell(v, i + 1).value = rapport3.cell(v, i).value

                                    rapport3.cell(v, i + 1).fill = PatternFill(patternType='solid',
                                                                              fgColor=self.my_yallow)
                                    kpid_instance.append(round(a, 2))



                                else:
                                    rapport3.cell(v, i + 1).value = rapport3.cell(v, i + 1).value
                                    rapport3.cell(v, i + 1).fill = PatternFill(patternType='solid', fgColor=self.my_yallow)
                                    kpid_instance.append(round(a,2))




                            else:

                                if b == 0 and str(rapport3.cell(v,
                                                               i - 4).value).strip() != "Only check \nrules eCLEM \nfill" and str(
                                        rapport3.cell(v, i - 4).value) != 'NA':
                                    print(rapport3.cell(v, i - 4).value)
                                    rapport3.cell(v, i + 1).value = 'NE'
                                    rapport3.cell(v, i + 1).fill = PatternFill(patternType='solid',
                                                                              fgColor=self.my_white)
                                    kpid_instance.append('NE')
                                elif str(rapport3.cell(v, i - 4).value) == "Only check \nrules eCLEM \nfill" or str(
                                        rapport3.cell(v, i - 4).value) != 'NA':
                                    rapport3.cell(v, i + 1).value = 'NA'
                                    rapport3.cell(v, i + 1).fill = PatternFill(patternType='solid', fgColor=self.my_gray)
                                    kpid_instance.append('NA')


                                else:
                                    rapport3.cell(v, i + 1).value = 'NA'
                                    rapport3.cell(v, i + 1).fill = PatternFill(patternType='solid', fgColor=self.my_gray)
                                    kpid_instance.append('NA')



                            #rapport.cell(v, i + 1).value = "hello"

                            rapport3.cell(v, i + 1).border = thin_border

                            kpidv_instance.append(v - 7)
                            cor='=Results_ASPICE!'+str(get_column_letter(i+1))+str(v)


                        rapport3.column_dimensions[get_column_letter(i + 1)].width = 15
                self.kpi_d.append(kpid_instance)
                self.kpi_dv.append(kpidv_instance)
        self.state2(self.rapport2_id, workbook, self.kpi_raport2, self.kpi_c, self.kpi_cv, self.kpi_d,self.kpi_dv)



    def insert_data(self,rapport3,workbook):# inserer les donneé de ID name , safty ,.. qui se trouve dans le fichier score

        self.idd=0

        self.rapport2_id=[]
        ##########################33

        header_formattxt = Alignment(wrap_text=True)

        for i in  range(len(self.list_report3)):
            x = 6
            if self.list_report3[i] in self.list_original and self.list_report3[i] not in [' ','Unnamed','Service']:
                b = []
                for m in list(self.data1[self.list_report3[i]]):

                    x += 1
                    if x<len( self.data1[self.list_report3[i]])+7:# pour supprimer les case qui contient la valeur nan dans letableau score
                        rapport3.cell(x, i + 1).value = m
                        b.append(m)
                        rapport3.cell(x, i+1).border = thin_border
                self.rapport2_id.append(b)

            if self.list_report3[i]=='Service':# inserer la colone service qui se trouve dans le fichier suivi ...
                b = []
                for m, n in zip(self.list_service, self.list_id):

                    x += 1
                    self.idd += 1
                    if n in list(self.data1['ID']):

                        v = list(self.data1['ID']).index(int(n)) + 7
                        rapport3.cell(v, i+1).value = m

                        #b.append(self.list_service[v-1])

                        b.append(v-7)
                        rapport3.cell(v, i+1).border = thin_border

                self.rapport2_id.append(b)

        self.state(rapport3 )
        self.insert_KPI(rapport3,self.workbook)

        #self.state2(self.rapport2_id,workbook,[])


        ####################################################################################################################







    def create_table(self): #creatio le tableau de rapport3 inserer les titres

        v=0
        list_jaune=[]
        list_bleu=[]
        self.read_data()

        for i in list_names_titles3:
            x=list_indice3[v]
            v+=1
            self.list_state_titles = ['Nb "G"\n'+x, 'Nb "O"\n'+x, 'Nb "R"\n'+x, 'Nb "NA"\n'+x, 'Nb "NE"\n'+x, '% G\n' + x,
                                  '% O\n' + x,
                                  '% R\n' + x, '% NA\n' + x, '% NE\n' + x, ' ', 'N\n'+x,
                                  'n.\n'+x,
                                  'KPI.1a\n' + x,
                                  'KPI.1b\n' + x, 'n-Nb"G"\n'+x, 'Nb. "NS_G"\n'+x, 'calcul KPI.1c\n'+x, 'KPI.1c\n' + x,
                                  'Nb "FS"\n'+x, 'Nb "PS"\n'+x, 'Nb "NS"\n'+x,
                                  'calcul KPI.1d\n'+x, 'KPI.1d\n' + x,
                                      ]
            list_jaune.extend([self.list_state_titles[i] for i in [13, 14, 18, 23]])
            list_bleu.extend([self.list_state_titles[i] for i in [5, 6, 7, 8, 9]])

        self.my_red = openpyxl.styles.colors.Color(rgb='00FF0000')
        self.my_green = openpyxl.styles.colors.Color(rgb='0066b615')
        self.my_orange = openpyxl.styles.colors.Color(rgb='00FFA500')
        self.my_gray = openpyxl.styles.colors.Color(rgb='00C0C0C0')
        self.my_bleu = openpyxl.styles.colors.Color(rgb='003399ff')
        self.my_yallow = openpyxl.styles.colors.Color(rgb='00FFFF00')
        my_blue = openpyxl.styles.colors.Color(rgb='001E90FF')
        self.my_black = openpyxl.styles.colors.Color(rgb='00000000')
        self.my_white = openpyxl.styles.colors.Color(rgb='00FFFFFF')
        header_formatfont = Font(bold=True, )
        header_formattxt = Alignment(wrap_text=True)


        rapport3 = self.workbook.create_sheet('Results_ASPICE')

        rapport3.title = 'Results_ASPICE'
        ref= 'A6:'+str(get_column_letter(len(self.list_report3)))+str(len(self.list_id)+6)


        tab = Table(displayName="Table2", ref=ref)
        # I list out the 4 show-xyz options here for reference
        style = TableStyleInfo(
            #name="TableStyleLight18",
            name="TableStyleLight21",
            #name="TableStyleMedium22",

            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
        tab.tableStyleInfo = style
        rapport3.add_table(tab)
        x=6
        all_names=[]
        for i in list_names_titles3:
            for j in i:
                all_names.append(j)

        for i in  range(len(list_report3)):
            rapport3.cell(x,i+1).value=self.list_report3[i]
            rapport3.cell(x, i+1).border = thin_border
            if rapport3.cell(x,i+1).value in list_head[0] :
                rapport3.cell(x,i+1).font = header_formatfont # font gras pour les titres
                rapport3.column_dimensions[get_column_letter(i + 1)].width = 15
            if rapport3.cell(x,i+1).value in self.list_red:rapport3.cell(x,i+1).font = Font(color="FFFF0000")# mette les coulleur rouge
            if rapport3.cell(x,i+1).value in list_jaune:
                rapport3.cell(x,i+1).fill = PatternFill(patternType='solid', fgColor=self.my_yallow)# mette les coulleur jaune
                rapport3.column_dimensions[get_column_letter(i + 1)].width = 15
            if rapport3.cell(x,i+1).value in list_bleu:
                rapport3.cell(x,i+1).fill = PatternFill(patternType='solid', fgColor=my_blue)# mette les coulleur bleu
                rapport3.column_dimensions[get_column_letter(i + 1)].width = 15

            rapport3.cell(x-1,i+1).fill = PatternFill(patternType='solid', fgColor=self.my_black)
            rapport3.cell(x,i+1).alignment = header_formattxt
            rapport3.column_dimensions[get_column_letter(i + 1)].width = 15
            if list_report3[i] not in list_head[0] and  list_report3[i] not in all_names:
                 rapport3.column_dimensions[get_column_letter(i + 1)].width = 15



        self.insert_data(rapport3,self.workbook)










        ##########################################################################################################################################################

    def state2(self, rap2_id, workbook, rap2_kpi_ab, kpi_c, kpi_cv, kpi_d, kpi_dv):  # raport4


        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))

        header_formatfont = Font(bold=True, )
        header_formattxt = Alignment(wrap_text=True)

        rapport = workbook.create_sheet('KPI.1_ASPICE')  # creation le deuxieme rapport

        a = 0

        list_head = [['ID', 'Name', 'Safety', 'MBD', 'SwQA', 'Direction', 'Service', 'ASWP deployment for'], ]
        list_kpi = ["KPI1.a", "KPI1.b", "KPI1.c", "KPI1.d"]
        list_kpi_dif = ["Deliverable Committment", "Deliverable Creation", "SWQA Check-Rate", "SWQA Status"]

        list_indice = list_indice3
        y = 0
        yy = 0
        z = 0

        for k in range(len(rap2_id[0])):
            for j in range(8):


                for i in range(7):
                    if i == 0:
                        if j<7:

                            c1 = get_column_letter(i + 2 + a) + str(j + 1)
                            c2 = get_column_letter(i + 7 + a) + str(j + 1)
                            c = c1 + ":" + c2

                            rapport.merge_cells(c)
                            rapport.column_dimensions[get_column_letter(i + 1 + a)].width = 40
                            rapport.merge_cells(c)
                        else:
                            c1 = get_column_letter(i + 2 + a) + str(j + 1)
                            c2 = get_column_letter(i + 5 + a) + str(j + 1)
                            c = c1 + ":" + c2
                            rapport.merge_cells(c)
                            rapport.column_dimensions[get_column_letter(i + 1 + a)].width = 40
                            rapport.merge_cells(c)
                            c1 = get_column_letter(i + 6 + a) + str(j + 1)
                            c2 = get_column_letter(i + 7 + a) + str(j + 1)
                            c = c1 + ":" + c2
                            rapport.merge_cells(c)
                            rapport.column_dimensions[get_column_letter(i + 1 + a)].width = 40
                            rapport.merge_cells(c)

                    else:
                        rapport.cell(j + 1, 1+a ).value = list_head[0][j]
                        rapport.cell(1, 2 + a).value = rap2_id[0][y]
                        rapport.cell(2, 2 + a).value = rap2_id[1][y]
                        rapport.cell(3, 2 + a).value = rap2_id[2][y]
                        rapport.cell(4, 2 + a).value = rap2_id[3][y]
                        rapport.cell(5, 2 + a).value = rap2_id[4][y]
                        rapport.cell(6, 2 + a).value = rap2_id[5][y]
                        rapport.cell(j + 10, 1 + a).value = rap2_id[1][y]
                        rapport.cell(j + 10, 1 + 2 + a).border = thin_border

                        rapport.cell(j + 1, 1+ a).border = thin_border

                        if rap2_id[0].index(rap2_id[0][y]) in rap2_id[6] :
                          rapport.cell(7, 2 + a).value = self.list_service[rap2_id[6].index(rap2_id[0].index(rap2_id[0][y]))]

                        rapport.cell(j + 1, i + 1 + a).border = thin_border

                        rapport.column_dimensions[get_column_letter(i + 1 + a)].width = 25
                        #rapport.cell(j + 1,  1 + a).font = Font(color="00FF8C00")
                        #rapport.cell(j + 1,  2 + a).font = Font(color="00FF8C00")
                        rapport.cell(j + 1, 1+a).font = Font(bold=True)


                    rapport.cell(j+1, i + 1 + a).alignment = Alignment(horizontal='center', vertical='center')



            y+=1
            x = 0
            x2=0
            x3=0

            for j in range(len(list_indice3)):


                for i in range(7):
                    if j == 0:
                        c1 = get_column_letter(1 + a) + str(8)
                        c2 = get_column_letter(1 + a) + str(9)

                        c = c1 + ":" + c2
                        rapport.merge_cells(c)


                        if i < 4:
                            rapport.cell(j + 9, i + 2 + a).value = list_kpi[i]
                            rapport.cell(j + 9, i + 2 + a).border = thin_border
                            rapport.cell(j + 9, i + 2 + a).alignment = Alignment(horizontal='center', vertical='center',
                                                                              )

                            rapport.cell(j + 10, i + 2 + a).value = list_kpi_dif[i]
                            rapport.cell(j + 10, i + 2 + a).border = thin_border
                            rapport.cell(j + 10, i + 2 + a).alignment = Alignment(horizontal='center', vertical='center',
                                                                              )


                            #rapport.cell(j + 8, i+2 + a).font = Font(color="00FF8C00")
                            rapport.cell(j + 10, i+2 + a).font = Font(bold=True,)
                            rapport.cell(j + 10,  1 + a).alignment = Alignment(horizontal='center', vertical='center',
                                                      )

                            rapport.cell(8,  2 + a).value = "KPI.1 Deliverable \nstatus"
                            rapport.cell(8,  2 + a).font = Font(bold=True,)
                            ##############
                            rapport.cell(8, 6 + a).value = "KPI.2 Process \napplication status"
                            rapport.cell(8, 6 + a).font = Font(bold=True, )
                        if i>4:
                            rapport.cell(j + 9, 7 + a).value = 'Template \napplication '
                            rapport.cell(j + 9, 6 + a).value ='Process \napplication '
                            rapport.cell(j + 9, 7 + a).font = Font(bold=True, )
                            rapport.cell(j + 9, 6 + a).font = Font(bold=True, )
                            rapport.cell(j + 9, 7 + a).border = thin_border
                            rapport.cell(j + 9, 6 + a).border = thin_border
                            rapport.cell(j + 9, 7 + a).alignment = Alignment(horizontal='center', vertical='center', )
                            rapport.cell(j + 9, 6 + a).alignment = Alignment(horizontal='center', vertical='center', )




                    if 0<i<3:
                        if len(rap2_kpi_ab) >0:

                            rapport.cell(j + 11, i + 1 + a).value = rap2_kpi_ab[z][x]
                            rapport.cell(j + 11, i + 1 + a).alignment = Alignment(horizontal='center',
                                                                                  vertical='center')
                            try:
                                valeur = float(str(rap2_kpi_ab[z][x]).replace('%', ''))
                                if valeur == 100:
                                    rapport.cell(j + 11, i + 1 + a).fill = PatternFill(patternType='solid',
                                                                                       fgColor=self.my_bleu)
                                if 75 <= valeur < 100:
                                    rapport.cell(j + 11, i + 1 + a).fill = PatternFill(patternType='solid',
                                                                                       fgColor=self.my_green)
                                if 50 <= valeur < 75:
                                    rapport.cell(j + 11, i + 1 + a).fill = PatternFill(patternType='solid',
                                                                                       fgColor=self.my_yallow)
                                if 25 <= valeur < 50:
                                    rapport.cell(j + 11, i + 1 + a).fill = PatternFill(patternType='solid',
                                                                                       fgColor=self.my_orange)
                                if 0 <= valeur < 25:
                                    rapport.cell(j + 11, i + 1 + a).fill = PatternFill(patternType='solid',
                                                                                       fgColor=self.my_red)
                            except ValueError:
                                valeur = str(rap2_kpi_ab[z][x]).replace('%', '')

                                if str(valeur)=='NA':
                                    rapport.cell(j + 11, i + 1 + a).fill = PatternFill(patternType='solid',fgColor=self.my_gray)
                            if str(valeur) == 'NA':
                                rapport.cell(j + 11, i + 1 + a).fill = PatternFill(patternType='solid',
                                                                                   fgColor=self.my_gray)

                            x += 1
                    if i==3:
                        if len(kpi_c) >0:
                            if rap2_id[0].index(rap2_id[0][yy]) in kpi_cv[x2]:
                                rapport.cell(j+11, i+1 + a).value = kpi_c[x2][ kpi_cv[x2].index(rap2_id[0].index(rap2_id[0][yy]))]




                                try:
                                    valeur = float(str(kpi_c[x2][ kpi_cv[x2].index(rap2_id[0].index(rap2_id[0][yy]))]).replace('%', ''))

                                    if valeur == 100:
                                        rapport.cell(j + 11, i + 1 + a).fill = PatternFill(patternType='solid',
                                                                                           fgColor=self.my_bleu)
                                    if 75 <= valeur < 100:
                                        rapport.cell(j + 11, i + 1 + a).fill = PatternFill(patternType='solid',
                                                                                           fgColor=self.my_green)
                                    if 50 <= valeur < 75:
                                        rapport.cell(j + 11, i + 1 + a).fill = PatternFill(patternType='solid',
                                                                                           fgColor=self.my_yallow)
                                    if 25 <= valeur < 50:
                                        rapport.cell(j + 11, i + 1 + a).fill = PatternFill(patternType='solid',
                                                                                           fgColor=self.my_orange)
                                    if 0 <= valeur < 25:
                                        rapport.cell(j + 11, i + 1 + a).fill = PatternFill(patternType='solid',
                                                                                           fgColor=self.my_red)
                                    if str(valeur) == 'NA':
                                        rapport.cell(j + 11, i + 1 + a).fill = PatternFill(patternType='solid',
                                                                                           fgColor=self.my_gray)

                                except ValueError:
                                    rapport.cell(j + 11, i + 1 + a).fill = PatternFill(patternType='solid',
                                                                                       fgColor=self.my_gray)

                            rapport.column_dimensions[get_column_letter(i + 1 + a)].width = 30
                            rapport.cell(j + 11, i + 1 + a).alignment = Alignment(horizontal='center',
                                                                                  vertical='center')
                            x2 += 1


                    if i==4:
                        if len(kpi_d) >0:
                            if rap2_id[0].index(rap2_id[0][yy]) in kpi_dv[x3]:
                                f = kpi_d[x3][kpi_dv[x3].index(rap2_id[0].index(rap2_id[0][yy]))]
                                if f != 'NA' and f != 'NE':
                                    f = str(f)+ '%'

                                rapport.cell(j + 11, i + 1 + a).value = f

                                try:
                                    valeur =float( str(kpi_d[x3][kpi_dv[x3].index(rap2_id[0].index(rap2_id[0][yy]))]).replace('%', ''))
                                    if valeur == 100:
                                        rapport.cell(j + 11, i + 1 + a).fill = PatternFill(patternType='solid',
                                                                                           fgColor=self.my_bleu)
                                    if 75 <= valeur < 100:
                                        rapport.cell(j + 11, i + 1 + a).fill = PatternFill(patternType='solid',
                                                                                           fgColor=self.my_green)
                                    if 50 <= valeur < 75:
                                        rapport.cell(j + 11, i + 1 + a).fill = PatternFill(patternType='solid',
                                                                                           fgColor=self.my_yallow)
                                    if 25 <= valeur < 50:
                                        rapport.cell(j + 11, i + 1 + a).fill = PatternFill(patternType='solid',
                                                                                           fgColor=self.my_orange)
                                    if 0 <= valeur < 25:
                                        rapport.cell(j + 11, i + 1 + a).fill = PatternFill(patternType='solid',
                                                                                       fgColor=self.my_red)

                                except   ValueError:
                                    if str(kpi_d[x3][kpi_dv[x3].index(rap2_id[0].index(rap2_id[0][yy]))]) == 'NA':
                                        rapport.cell(j + 11, i + 1 + a).fill = PatternFill(patternType='solid',
                                                                                           fgColor=self.my_gray)
                                    else:
                                        rapport.cell(j + 11, i + 1 + a).fill = PatternFill(patternType='solid',
                                                                                           fgColor=self.my_white)



                            x3+=1
                            rapport.column_dimensions[get_column_letter(i + 1 + a)].width = 20
                            rapport.cell(j + 11, i + 1 + a).alignment = Alignment(horizontal='center',
                                                                                  vertical='center')



                    if i>4:
                        c1 = get_column_letter(6 + a) + str(9)
                        c2 = get_column_letter(6 + a) + str(10)

                        c = c1 + ":" + c2
                        rapport.merge_cells(c)
                        c1 = get_column_letter(7 + a) + str(9)
                        c2 = get_column_letter(7 + a) + str(10)

                        c = c1 + ":" + c2
                        rapport.merge_cells(c)



                        rapport.cell(j + 11, i + 1 + a).border = thin_border
                        rapport.cell(j + 11, i + 1 + a).fill = PatternFill(patternType='solid',
                                                                           fgColor=self.my_gray)

                        rapport.column_dimensions[get_column_letter(i + 1 + a)].width = 20



                    rapport.cell(j + 11, 1 + a).value = list_indice[j]
                    rapport.cell(j + 11, i + 1 + a).border = thin_border


                   # rapport.cell(j + 11, i + 1 + a).alignment = Alignment(horizontal='center', vertical='center')
                    #rapport.cell(j+11,1+a).fill = PatternFill(patternType='solid')
                    rapport.cell(j + 11, 1+ a).font = Font(bold=True, )


            a += 8
            z+=1
            yy += 1
        img = openpyxl.drawing.image.Image('fig.png')

        img.width = 600
        img.height = 250
        rapport.add_image(img, 'B31')





    def state(self, rapport3):
        self.list_id_score = list(self.data1['ID'])
        self.list_mbd_score = list(self.data1['MBD'])
        self.list_SwQA_score = list(self.data1['SwQA'])


        def traiter(args):

            i_itmes = []
            for i in range(len(list_names_titles3)):
                i_itmes.append(i)  # pour vérifier est ce que self.valeur est compatible
            w = 6
            a = 0
            self.kpi_raport2 = []
            kpi_instance = []

            for k in range(len(self.list_id_net)):


                w += 1
                s = -1
                kpi_instance = []
                for arg in args:



                    gg = 0
                    oo = 0
                    rr = 0
                    nna = 0
                    nne = 0
                    pper = 0

                    fglobal = '='
                    s += 1
                    arg.sort()

                    ffs = 0
                    pps = 0
                    nns = 0
                    nns_g = 0
                    NG=0

                    for i, j in zip(i_itmes, arg):


                        if i in self.valeur:
                            g = 0
                            o = 0
                            r = 0
                            na = 0
                            ne = 0
                            per = 0
                            fs = 0
                            ps = 0
                            ns = 0
                            ns_g = 0
                            N=0

                            for n in self.list_red:
                                if n in list_names_titles3[i]:
                                    if n in list_id_FullSW:

                                        index=self.list_id_score.index(self.list_id_net[k])
                                        if self.list_SwQA_score[index].upper()=='FULL SW' and  self.list_mbd_score[index].upper() =='YES':



                                            N+=list_yes_FullSW[list_id_FullSW.index(n)]
                                            NG+=list_yes_FullSW[list_id_FullSW.index(n)]

                                        if self.list_SwQA_score[index].upper()=='FULL SW' and  self.list_mbd_score[index].upper() =='NO':
                                             N+=list_no_FullSW[list_id_FullSW.index(n)]
                                             NG += list_no_FullSW[list_id_FullSW.index(n)]
                                        if self.list_SwQA_score[index].upper()=='PA':
                                             N+=list_pa_FullSW[list_id_FullSW.index(n)]
                                             NG += list_pa_FullSW[list_id_FullSW.index(n)]


                                    if self.list_id_net[k] in list(self.data3['ID CLEM O52_ASWP']):
                                        f=self.data3[n][list(self.data3['ID CLEM O52_ASWP']).index(self.list_id_net[k])]
                                        if str(f).upper()=='FS':
                                            fs+=1
                                            ffs+=1
                                        if str(f).upper()=='PS':
                                            ps+=1
                                            pps+=1

                                        if str(f).upper()=='NS':
                                            ns+=1
                                            nns+=1
                                        if str(f).upper()=='NS_G':
                                            ns_g+=1
                                            nns_g+=1

                                    per += 1
                                    pper += 1
                                    instance = list(self.data1[n])[k]

                                    if str(instance).strip() == 'G':
                                        g += 1
                                        gg += 1
                                        rapport3.cell(w, self.list_report3.index(n) + 1).fill = PatternFill(
                                            patternType='solid', fgColor=self.my_green)

                                    if str(instance).strip() == 'O':
                                        o += 1
                                        oo += 1
                                        rapport3.cell(w, self.list_report3.index(n) + 1).fill = PatternFill(
                                            patternType='solid', fgColor=self.my_orange)
                                    if str(instance).strip() == 'R':
                                        r += 1
                                        rr += 1

                                        rapport3.cell(w, self.list_report3.index(n) + 1).fill = PatternFill(
                                            patternType='solid', fgColor=self.my_red)
                                    if str(instance).strip() == '':
                                        rapport3.cell(w, self.list_report3.index(n) + 1).value = 'NO DATA'

                                    if str(instance).strip() == 'NA':
                                        na += 1
                                        nna += 1
                                        rapport3.cell(w, self.list_report3.index(n) + 1).value = 'NA'
                                        rapport3.cell(w, self.list_report3.index(n) + 1).fill = PatternFill(
                                            patternType='solid', fgColor=self.my_gray)

                                    if str(instance).strip() == 'NE':
                                        rapport3.cell(w, self.list_report3.index(n) + 1).fill = PatternFill(
                                            patternType='solid', fgColor=self.my_white)
                                        ne += 1
                                        nne += 1

                            ################################lie a tout les pertie qui des titre rouge###############################
                            go = [g, o, r, na, ne]
                            fo = [fs, ps, ns, ns_g]
                            if sum(go)==0:
                                rapport3.cell(w, j + 1).value = 'NO DATA'
                                rapport3.cell(w, j + 6).value = 'NO DATA'
                                rapport3.cell(w, j + 8).value = 'NO DATA'  # N
                                rapport3.cell(w, j + 9).value = 'NO DATA'  # n
                                rapport3.cell(w, j + 10).value = 'NO DATA'  # n
                                rapport3.cell(w, j + 11).value = 'NO DATA'  # n

                                kpi_instance.append('NO DATA')

                                rapport3.cell(w, j + 12).value = 'NO DATA' # n-Nb"G"
                                rapport3.cell(w, j + 13).value = 'NO DATA' # n_NB"NS_G"
                                rapport3.cell(w, j + 14).value = 'NO DATA'
                                rapport3.cell(w, j + 16).value = 'NO DATA'  # FS########
                                rapport3.cell(w, j + 17).value = 'NO DATA'  # FS########
                                rapport3.cell(w, j + 18).value = 'NO DATA'  # FS########
                                rapport3.cell(w, j + 19).value = 'NO DATA'  # FS########
                                if s == len(go) - 1:

                                    if a < len(list_indice3):
                                        rapport3.cell(4, j + 15).value = list_indice3[
                                            a]  # mette le titre des partie des tableau  a partir de list list_indice
                                        rapport3.cell(4, j + 15).font = Font(bold=True, size=18)
                                        rapport3.cell(4, j - 3).value = list_indice3[
                                            a]  # mette le titre des partie des tableau  a partir de list list_indice
                                        rapport3.cell(4, j - 3).font = Font(bold=True, size=18)

                                        a += 1



                            else:



                                    rapport3.cell(w, j + 1).value = go[s]

                                    fx1 = '=NB.SI(' + str(get_column_letter(j - 5)) + str(w) + ':' + str(
                                        get_column_letter(j)) + str(w) + ',"O")'
                                    # fx2='=LEN( '+str(get_column_letter(j -5))+str(w)+')-LEN(SUBSTITUTE('+ str(get_column_letter(j -5))+str(w)+',"z",""))'
                                    # rapport.cell(w, j + 1).value = fx2

                                    # rapport.cell(w, j + 1).value = '= NB.SI.ENS(I7:N7;"O")'

                                    # rapport.cell(w, j + 6).value = str(round(100*go[s]/per,2))+"%"
                                    fx6 = '=' + str(get_column_letter(j + 1)) + str(w) + '/' + str(per)

                                    # rapport.cell(w, j + 6).value = fx6
                                    rapport3.cell(w, j + 6).value = str(round((100 * go[s] / per), 2)) + "%"
                                    rapport3.cell(w, j + 1).border = thin_border
                                    rapport3.cell(w, j + 6).border = thin_border
                                    if s == len(go) - 1:

                                        if a < len(list_indice3):
                                            rapport3.cell(4, j + 15).value = list_indice3[
                                                a]  # mette le titre des partie des tableau  a partir de list list_indice
                                            rapport3.cell(4, j + 15).font = Font(bold=True, size=18)
                                            rapport3.cell(4, j - 3).value = list_indice3[
                                                a]  # mette le titre des partie des tableau  a partir de list list_indice
                                            rapport3.cell(4, j - 3).font = Font(bold=True, size=18)

                                            a += 1


                                        rapport3.cell(w, j + 8).border = thin_border
                                        rapport3.cell(w, j + 9).border = thin_border
                                        rapport3.cell(w - 6, j + 7).fill = PatternFill(patternType='solid',
                                                                                       fgColor=self.my_gray)
                                        rapport3.cell(w, j + 7).fill = PatternFill(patternType='solid', fgColor=self.my_gray)
                                        rapport3.cell(w - 6, j + 21).fill = PatternFill(patternType='solid',
                                                                                        fgColor=self.my_black)  # le vide noir entre les changement
                                        rapport3.cell(w, j + 21).fill = PatternFill(patternType='solid',
                                                                                    fgColor=self.my_black)  # le vide noir entre les changement
                                        rapport3.cell(w, j + 8).value = N # N
                                        rapport3.cell(w, j + 9).value = N  # n

                                        if N > 0:
                                            rapport3.cell(w, j + 10).value = str(
                                                per / per * 100) + '%'  # KPI.1a Project Management###########################################################

                                            rapport3.cell(w, j + 11).value = str(
                                                round(((go[0] * 1 + go[1] * 0.5 + go[2] * 0) / per * 100),
                                                      2)) + '%'  # KPI.1bProject Management
                                            kpi_instance.append(str(per / per * 100) + '%')
                                            kpi_instance.append(str(
                                                round(((go[0] * 1 + go[1] * 0.5 + go[2] * 0) / per * 100),
                                                      2)) + '%' )

                                        else:
                                            rapport3.cell(w, j + 10).value = 'NA'

                                            rapport3.cell(w, j + 11).value = 'NA'
                                            kpi_instance.append('NA')
                                            kpi_instance.append('NA')




                                        rapport3.cell(w, j + 12).value = go[0]  # n-Nb"G"
                                        rapport3.cell(w, j + 13).value = fo[3]  # n_NB"NS_G"

                                        try:
                                            rapport3.cell(w, j + 14).value = str(
                                                round((go[0] - fo[3]) * 100 / go[0])) + '%'
                                        except ZeroDivisionError as error:
                                            rapport3.cell(w, j + 14).value = '0.00%'


                                        rapport3.cell(w, j + 16).value = fo[0]  # FS########
                                        rapport3.cell(w, j + 17).value = fo[1]
                                        rapport3.cell(w, j + 18).value = fo[2]

                                        # fonction pour automatiquer de calcule
                                        fx19 = '(' + str(get_column_letter(j + 16)) + str(w) + '*1+' + str(
                                            get_column_letter(j + 17)) + str(w) + '*0.5+' + str(
                                            get_column_letter(j + 18)) + str(w) + '*0)*100/' + "(" + str(
                                            get_column_letter(j + 12)) + str(w) + '-' + str(get_column_letter(j + 13)) + str(
                                            w) + ")"  # # pour  automatiser
                                        fxglo = '(' + str(get_column_letter(j + 16)) + str(w) + '*1+' + str(
                                            get_column_letter(j + 17)) + str(w) + '*0.5+' + str(
                                            get_column_letter(j + 18)) + str(w) + '*0)*100/' + str(
                                            get_column_letter(j + 13)) + str(w) + '+'  # pour  automatiser
                                        fglobal += fxglo
                                        # rapport.cell(w, j + 19).value = '= (BH7*1 + BI7*0,5 + BJ7*0) /BE7'
                                        # rapport.cell(w, j + 19).value = fx19

                                        if go[0]-fo[3]>0:
                                            rapport3.cell(w,j + 19).value = '=IF(' + fx19 + ' > 0," "' + '& ' + fx19 + ' & "%",' + fx19 + ' & "%")' # clacule KPI.1D
                                        else:
                                            rapport3.cell(w, j + 19).value ='0.00%'


                        ########################################partie global lie a mist gooo############################################
                        elif i == i_itmes[-1]:  # i est l index de  dernier parier de GLOBAL
                            goo = [gg, oo, rr, nna, nne]
                            foo = [ffs, pps, nns, nns_g]
                            if sum(goo)==0:
                                rapport3.cell(w, j + 1).value = 'NO DATA'
                                rapport3.cell(w, j + 6).value = 'NO DATA'
                                rapport3.cell(w, j + 8).value = 'NO DATA'  # N
                                rapport3.cell(w, j + 9).value = 'NO DATA'  # n
                                rapport3.cell(w, j + 10).value = 'NO DATA'  # n
                                rapport3.cell(w, j + 11).value = 'NO DATA'  # n

                                kpi_instance.append('NO DATA')

                                rapport3.cell(w, j + 12).value = 'NO DATA' # n-Nb"G"
                                rapport3.cell(w, j + 13).value = 'NO DATA' # n_NB"NS_G"
                                rapport3.cell(w, j + 14).value = 'NO DATA'
                                rapport3.cell(w, j + 16).value = 'NO DATA'  # FS########
                                rapport3.cell(w, j + 17).value = 'NO DATA'  # FS########
                                rapport3.cell(w, j + 18).value = 'NO DATA'  # FS########
                                rapport3.cell(w, j + 19).value = 'NO DATA'  # FS########
                                if s == len(goo) - 1:

                                    if a < len(list_indice3):
                                        rapport3.cell(4, j + 15).value = list_indice3[
                                            a]  # mette le titre des partie des tableau  a partir de list list_indice
                                        rapport3.cell(4, j + 15).font = Font(bold=True, size=18)
                                        rapport3.cell(4, j - 3).value = list_indice3[
                                            a]  # mette le titre des partie des tableau  a partir de list list_indice
                                        rapport3.cell(4, j - 3).font = Font(bold=True, size=18)

                                        a += 1
                            else:
                                rapport3.cell(w, j + 1).value = goo[s]
                                fx6 = '=' + str(get_column_letter(j + 1)) + str(w) + '/' + str(pper) + '*100 '

                                #rapport3.cell(w, j + 6).value = fx6
                                rapport3.cell(w, j + 6).value = str(round((100 * goo[s] / pper), 2)) + "%"
                                rapport3.cell(w, j + 1).border = thin_border
                                rapport3.cell(w, j + 6).border = thin_border
                                if s == len(goo) - 1:

                                    if a < len(list_indice3):
                                        rapport3.cell(4, j + 15).value = list_indice3[
                                            a]  # mette le titre des partie des tableau  a partir de list list_indice
                                        rapport3.cell(4, j + 15).font = Font(bold=True, size=18)
                                        rapport3.cell(4, j - 3).value = list_indice3[
                                            a]  # mette le titre des partie des tableau  a partir de list list_indice
                                        rapport3.cell(4, j - 3).font = Font(bold=True, size=18)

                                        a += 1
                                    rapport3.cell(w, j + 8).border = thin_border
                                    rapport3.cell(w, j + 9).border = thin_border
                                    rapport3.cell(w - 6, j + 7).fill = PatternFill(patternType='solid',
                                                                                   fgColor=self.my_gray)
                                    rapport3.cell(w, j + 7).fill = PatternFill(patternType='solid', fgColor=self.my_gray)
                                    rapport3.cell(w - 6, j + 21).fill = PatternFill(patternType='solid',
                                                                                    fgColor=self.my_black)  # le vide noir entre les changement
                                    rapport3.cell(w, j + 21).fill = PatternFill(patternType='solid',
                                                                                fgColor=self.my_black)  # le vide noir entre les changement
                                    rapport3.cell(w, j + 8).value = NG # N
                                    rapport3.cell(w, j + 9).value = NG  # n
                                    if NG> 0:
                                        rapport3.cell(w, j + 10).value = str(
                                            pper / pper * 100) + '%'  # KPI.1a Project Management
                                        rapport3.cell(w, j + 11).value = str(
                                            round(((goo[0] * 1 + goo[1] * 0.5 + goo[2] * 0) / pper * 100),
                                                  2)) + '%'  # KPI.1bProject Management
                                        kpi_instance.append(str(pper / pper * 100) + '%')
                                        kpi_instance.append(str(
                                            round(((goo[0] * 1 + goo[1] * 0.5 + goo[2] * 0) / pper * 100),
                                                  2)) + '%' )
                                    else:
                                        rapport3.cell(w, j + 10).value = 'NA'

                                        rapport3.cell(w, j + 11).value = 'NA'
                                        kpi_instance.append("NA")
                                        kpi_instance.append("NA")





                                    rapport3.cell(w, j + 12).value = goo[0]  # n-Nb"G"
                                    rapport3.cell(w, j + 13).value = foo[3]  # n_NB"NS_G"
                                    try:
                                        rapport3.cell(w, j + 14).value = str(
                                            round((goo[0]-foo[3])*100/goo[0] )) + '%' # calcule KPI.1c
                                    except ZeroDivisionError as error:
                                        rapport3.cell(w, j + 14).value ='NA'


                                    rapport3.cell(w, j + 16).value = foo[0]  # FS########
                                    rapport3.cell(w, j + 17).value = foo[1]
                                    rapport3.cell(w, j + 18).value = foo[2]
                                    # fonction pour automatiquer de calcule
                                    fx19 = '=(' + str(get_column_letter(j + 16)) + str(w) + '*1+' + str(
                                        get_column_letter(j + 17)) + str(w) + '*0.5+' + str(
                                        get_column_letter(j + 18)) + str(w) + '*0)/' + str(
                                        get_column_letter(j + 13)) + str(
                                        w)  # pour  automatiser
                                    # rapport.cell(w, j + 19).value = '= (BH7*1 + BI7*0,5 + BJ7*0) /BE7'
                                    # =SI((B2 - A2) / A2 % > 0;"+ " & (B2 - A2) / A2 % & " %";(B2 - A2) / A2 % & " %")
                                    # ="=SI("+fxglo2 +" > 0;"+ " & "+fxglo2+" &  '%';"+fxglo2+"& " '%'")



                                    fxglo2 = '(' + str(get_column_letter(j + 16)) + str(w) + '*1+' + str(
                                        get_column_letter(j + 17)) + str(w) + '*0.5+' + str(
                                        get_column_letter(j + 18)) + str(w) + '*0)*100/' + "(" + str(
                                        get_column_letter(j + 12)) + str(
                                        w) + '-' + str(
                                        get_column_letter(j + 13)) + str(
                                        w) + ")"

                                    # rapport.cell(w, j + 19).value = 'IF(+'+str(fxglo2)+'  > 0,'+""+str(fxglo2)+"'%'"+',"False")'
                                    if goo[0]-foo[3] >0:

                                        rapport3.cell(w,
                                                     j + 19).value = '=IF(' + fxglo2 + ' > 0," "' + '& ' + fxglo2 + ' & "%",' + fxglo2 + ' & "%")'  # calcule KPI.1D
                                    else:
                                        rapport3.cell(w, j + 19).value ='0.00%'




                        ###############################################parite lie a les elemnt qui ont aps aucun titre rouge
                        else:

                            go = [1, 2, 3, 4, 5]
                            rapport3.cell(w, j + 1).value = 0
                            rapport3.cell(w, j + 6).value = "0%"
                            rapport3.cell(w, j + 1).border = thin_border
                            rapport3.cell(w, j + 6).border = thin_border
                            if s == len(go) - 1:
                                if a < len(list_indice3):
                                    rapport3.cell(4, j + 15).value = list_indice3[
                                        a]  # mette le titre des partie des tableau  a partir de list list_indice
                                    rapport3.cell(4, j + 15).font = Font(bold=True, size=18)
                                    rapport3.cell(4, j - 3).value = list_indice3[
                                        a]  # mette le titre des partie des tableau  a partir de list list_indice
                                    rapport3.cell(4, j - 3).font = Font(bold=True, size=18)

                                    a += 1
                                rapport3.cell(w, j + 8).border = thin_border
                                rapport3.cell(w, j + 9).border = thin_border
                                rapport3.cell(w - 6, j + 7).fill = PatternFill(patternType='solid',
                                                                               fgColor=self.my_gray)  # c'est pour siparation en gray '
                                rapport3.cell(w, j + 7).fill = PatternFill(patternType='solid', fgColor=self.my_gray)
                                rapport3.cell(w - 6, j + 21).fill = PatternFill(patternType='solid',
                                                                                fgColor=self.my_black)
                                rapport3.cell(w, j + 21).fill = PatternFill(patternType='solid',
                                                                            fgColor=self.my_black)
                                rapport3.cell(w, j + 8).value = 0  # N
                                rapport3.cell(w, j + 9).value = 0  # n
                                rapport3.cell(w, j + 10).value = 'NA'  # KPI.1a Project Management
                                rapport3.cell(w, j + 11).value = 'NA'
                                kpi_instance.append('NA')# pour l'envoyer
                                kpi_instance.append('NA')

                                rapport3.cell(w, j + 12).value = 0  # n-Nb"R"
                                rapport3.cell(w, j + 13).value = 0  # V
                                if go[2] != 4:
                                    rapport3.cell(w, j + 14).value = '0%'
                                else:
                                    rapport3.cell(w, j + 14).value = '0' + '%'

                self.kpi_raport2.append(kpi_instance)




        self.valeur = []
        self.list_c2 = []  # list des index des coloone de state des couleur
        v = 0
        n = []
        g = []
        o = []
        r = []
        na = []

        for i in list_names_titles3:
            x = list_indice3[v]
            v += 1

            g.append('Nb "G"\n' + x)
            o.append('Nb "O"\n' + x)
            r.append('Nb "R"\n' + x)
            na.append('Nb "NA"\n' + x)
            n.append('Nb "NE"\n' + x)
        n_index = []
        g_index = []
        o_index = []
        r_index = []
        na_index = []
        for k in g:
            g_index.append(self.list_report3.index(k))
        for k in n:
            n_index.append(self.list_report3.index(k))

        for k in o:
            o_index.append(self.list_report3.index(k))
        for k in r:
            r_index.append(self.list_report3.index(k))
        for k in na:
            na_index.append(self.list_report3.index(k))


        list_c = [g_index, o_index, r_index, na_index, n_index]


        for i in self.list_report3:

            for j in list_names_titles3:


                if i in j and i in self.list_red:

                    if list_names_titles3.index(j) not in self.valeur:

                        self.valeur.append(list_names_titles3.index(j))
                    for p in list_c:

                        self.z = []
                        if p in list_c:

                            for h in range(len(p)):
                                if p[h] not in self.z:
                                    self.z.append(p[h])
                            if self.z not in self.list_c2:
                                self.list_c2.append(self.z)
                        else:

                            for h in range(len(p)):
                                if h % 2 == 0:
                                    if p[h] not in self.z:
                                        self.z.append(p[h])
                            if self.z not in self.list_c2:
                                self.list_c2.append(self.z)


        w = 6

        traiter(list_c)



